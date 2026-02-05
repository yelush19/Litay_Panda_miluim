#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
המרת מערכת ישנה לתבנית חדשה
מעתיק: עובדים, תקופות, תשלומי ב"ל
"""

from openpyxl import load_workbook
from datetime import datetime
import shutil
import os

OLD_FILE = r"C:\Projects\LitayPandaMiluim\מערכת_מילואים_מלאה.xlsx"
NEW_TEMPLATE = r"C:\Projects\LitayPandaMiluim\מערכת_מילואים_תבנית_חדשה.xlsx"
OUTPUT_FILE = r"C:\Projects\LitayPandaMiluim\מערכת_מילואים_מומרת.xlsx"

def convert():
    """המרה מקובץ ישן לחדש"""
    
    print("=" * 60)
    print("🔄 המרת מערכת ישנה לתבנית חדשה")
    print("=" * 60)
    
    # טעינה
    print("\n📂 טוען קבצים...")
    wb_old = load_workbook(OLD_FILE)
    wb_new = load_workbook(NEW_TEMPLATE)
    
    # 1. העתקת עובדים
    print("\n👥 מעתיק רשימת עובדים...")
    ws_old_emp = wb_old['1️⃣ רשימת עובדים']
    ws_new_emp = wb_new['1️⃣ רשימת עובדים']
    
    emp_count = 0
    for row in range(2, ws_old_emp.max_row + 1):
        emp_count += 1
        for col in range(1, 11):  # 10 עמודות
            ws_new_emp.cell(row, col).value = ws_old_emp.cell(row, col).value
    
    print(f"   ✅ {emp_count} עובדים הועתקו")
    
    # 2. העתקת תקופות מילואים
    print("\n📋 מעתיק תקופות מילואים...")
    ws_old_periods = wb_old['2️⃣ תקופות מילואים']
    ws_new_track = wb_new['📊 מעקב מילואים ותשלומים']
    
    # מיפוי עמודות ישן → חדש
    # ישן: מזהה, שם, מחלקה, התחלה, סיום, חודש, ימים, א-ה, שישי, שבת, חג, תעריף, תשלום...
    # חדש: אותו דבר + עמודות נוספות
    
    period_count = 0
    for row in range(2, ws_old_periods.max_row + 1):
        period_count += 1
        
        # העתקת 18 עמודות ראשונות
        for col in range(1, 19):
            ws_new_track.cell(row, col).value = ws_old_periods.cell(row, col).value
        
        # עמודה 19: מועד תשלום ביטוח לאומי (מעמודה 16 בישן)
        ws_new_track.cell(row, 18).value = ws_old_periods.cell(row, 16).value
        
        # עמודה 20: חודש ביצוע תשלום (מעמודה 18 בישן)
        ws_new_track.cell(row, 20).value = ws_old_periods.cell(row, 18).value
        
        # עמודה 22: הערות (מעמודה 19 בישן)
        ws_new_track.cell(row, 22).value = ws_old_periods.cell(row, 19).value
    
    print(f"   ✅ {period_count} תקופות הועתקו")
    
    # 3. העתקת תשלומי ב"ל
    print("\n💰 מעתיק תשלומי ב\"ל...")
    ws_old_btl = wb_old['3️⃣ תשלומי ב"ל']
    ws_new_btl = wb_new['3️⃣ תשלומי ב"ל']
    
    btl_count = 0
    for row in range(2, ws_old_btl.max_row + 1):
        btl_count += 1
        for col in range(1, 13):  # 12 עמודות
            ws_new_btl.cell(row, col).value = ws_old_btl.cell(row, col).value
    
    print(f"   ✅ {btl_count} תשלומים הועתקו")
    
    # 4. העתקת רשימת תשלומים
    print("\n💵 מעתיק רשימת תשלומים...")
    ws_old_pay = wb_old['💵 רשימת תשלומים']
    ws_new_pay = wb_new['💵 רשימת תשלומים']
    
    pay_count = 0
    for row in range(2, ws_old_pay.max_row + 1):
        pay_count += 1
        for col in range(1, 7):  # 6 עמודות
            ws_new_pay.cell(row, col).value = ws_old_pay.cell(row, col).value
    
    print(f"   ✅ {pay_count} מנות הועתקו")
    
    # שמירה
    wb_new.save(OUTPUT_FILE)
    wb_old.close()
    wb_new.close()
    
    print("\n" + "=" * 60)
    print("✅ המרה הושלמה בהצלחה!")
    print("=" * 60)
    print(f"\n📁 קובץ חדש נשמר ב:\n   {OUTPUT_FILE}")
    print(f"\n📊 סיכום:")
    print(f"   👥 עובדים: {emp_count}")
    print(f"   📋 תקופות: {period_count}")
    print(f"   💰 תשלומי ב\"ל: {btl_count}")
    print(f"   💵 מנות: {pay_count}")
    
    input("\nלחץ Enter לסגירה...")

if __name__ == "__main__":
    try:
        if not os.path.exists(OLD_FILE):
            print(f"❌ לא נמצא קובץ ישן:\n   {OLD_FILE}")
            input("\nלחץ Enter לסגירה...")
        elif not os.path.exists(NEW_TEMPLATE):
            print(f"❌ לא נמצאה תבנית חדשה:\n   {NEW_TEMPLATE}")
            input("\nלחץ Enter לסגירה...")
        else:
            convert()
    except Exception as e:
        print(f"\n❌ שגיאה: {e}")
        input("\nלחץ Enter לסגירה...")
