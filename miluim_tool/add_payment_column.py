#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
הוספת עמודה: סכום ששולם בפועל לעובד
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import shutil
import os

SYSTEM_FILE = r"C:\Projects\LitayPandaMiluim\מערכת_מילואים_מלאה.xlsx"

# צבעי ליטאי
GREEN_HEADER = PatternFill(start_color="528163", end_color="528163", fill_type="solid")
GREEN_LIGHT = PatternFill(start_color="8dd1bb", end_color="8dd1bb", fill_type="solid")
header_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")

def add_payment_column():
    """הוספת עמודת תשלום בפועל"""
    
    print("=" * 60)
    print("➕ הוספת עמודה: סכום ששולם בפועל לעובד")
    print("=" * 60)
    
    if not os.path.exists(SYSTEM_FILE):
        print(f"\n❌ לא נמצא קובץ: {SYSTEM_FILE}")
        input("\nלחץ Enter לסגירה...")
        return
    
    # גיבוי
    backup_dir = os.path.join(os.path.dirname(SYSTEM_FILE), "backups")
    os.makedirs(backup_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = os.path.join(backup_dir, f"backup_before_add_column_{timestamp}.xlsx")
    shutil.copy2(SYSTEM_FILE, backup_path)
    print(f"\n💾 גיבוי: {os.path.basename(backup_path)}")
    
    # טעינה
    wb = load_workbook(SYSTEM_FILE)
    
    # זיהוי שם הגיליון
    if '📊 מעקב מילואים ותשלומים' in wb.sheetnames:
        sheet_name = '📊 מעקב מילואים ותשלומים'
    elif '2️⃣ תקופות מילואים' in wb.sheetnames:
        sheet_name = '2️⃣ תקופות מילואים'
    else:
        print("\n❌ לא נמצא גיליון מעקב!")
        wb.close()
        input("\nלחץ Enter לסגירה...")
        return
    
    ws = wb[sheet_name]
    
    print(f"\n📄 גיליון: {sheet_name}")
    print(f"   עמודות נוכחיות: {ws.max_column}")
    
    # בדיקה אם העמודה כבר קיימת
    if ws.max_column >= 21:
        col_21_header = ws.cell(1, 21).value
        if col_21_header and "שולם" in str(col_21_header):
            print(f"\n✅ עמודה 21 כבר קיימת: {col_21_header}")
            wb.close()
            input("\nלחץ Enter לסגירה...")
            return
    
    # הוספת כותרת בעמודה 21
    new_col = 21
    ws.cell(1, new_col).value = "💰 סכום ששולם בפועל לעובד"
    ws.cell(1, new_col).font = header_font
    ws.cell(1, new_col).fill = GREEN_HEADER
    ws.cell(1, new_col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # הגדרת רוחב עמודה
    ws.column_dimensions[chr(64 + new_col)].width = 18
    
    # צביעת תאים בשורות קיימות (ירוק בהיר)
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row, new_col)
        cell.fill = GREEN_LIGHT
        cell.alignment = Alignment(horizontal='right', vertical='center')
    
    print(f"\n✅ נוספה עמודה {new_col}:")
    print(f"   כותרת: 💰 סכום ששולם בפועל לעובד")
    print(f"   רוחב: 18")
    print(f"   צבע: ירוק בהיר (ליטאי)")
    
    # עדכון מבנה הגיליון
    print(f"\n📊 מבנה מעודכן ({ws.max_column} עמודות):")
    for col in range(1, ws.max_column + 1):
        header = ws.cell(1, col).value
        marker = "✨" if col == new_col else "  "
        print(f"   {marker} {col}. {header}")
    
    # שמירה
    wb.save(SYSTEM_FILE)
    wb.close()
    
    print("\n" + "=" * 60)
    print("✅ עמודה נוספה בהצלחה!")
    print("=" * 60)
    
    print("\n📝 איך להשתמש:")
    print("   1. עמודה 18 = סכום הפרשים (מחושב)")
    print("   2. עמודה 19 = חודש ביצוע תשלום")
    print("   3. עמודה 20 = הערות")
    print("   4. עמודה 21 = סכום ששולם בפועל ← הזן ידנית!")
    
    print("\n🎯 דוגמה לסינון:")
    print("   • סנן לפי חודש ביצוע = 05/2025")
    print("   • סכום עמודה 18 = כמה צריך לשלם")
    print("   • סכום עמודה 21 = כמה שולם בפועל")
    print("   • הפרש = עמודה 21 - עמודה 18")
    
    input("\nלחץ Enter לסגירה...")

if __name__ == "__main__":
    try:
        add_payment_column()
    except Exception as e:
        print(f"\n❌ שגיאה: {e}")
        import traceback
        traceback.print_exc()
        input("\nלחץ Enter לסגירה...")
