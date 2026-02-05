#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ניקוי חלקי של המערכת
שומר: תקופות מילואים + תשלומי ב"ל (לפני 28/11/2025)
מוחק: דוח מסכם + רשימת תשלומים (יחושבו מחדש)
"""

from openpyxl import load_workbook
from datetime import datetime
import shutil
import os

SYSTEM_FILE = r"C:\Projects\LitayPandaMiluim\מערכת_מילואים_מלאה.xlsx"

def backup_file():
    """יצירת גיבוי"""
    if os.path.exists(SYSTEM_FILE):
        backup_dir = os.path.join(os.path.dirname(SYSTEM_FILE), "backups")
        os.makedirs(backup_dir, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = os.path.join(backup_dir, f"backup_before_clean_{timestamp}.xlsx")
        shutil.copy2(SYSTEM_FILE, backup_path)
        print(f"✅ גיבוי נוצר: {backup_path}")
        return backup_path
    return None

def clean_summary_and_payments():
    """מחיקת דוח מסכם ורשימת תשלומים בלבד"""
    
    print("=" * 60)
    print("🧹 ניקוי חלקי של המערכת")
    print("=" * 60)
    
    # גיבוי
    backup_path = backup_file()
    
    # טעינת הקובץ
    wb = load_workbook(SYSTEM_FILE)
    
    # 1. בדיקת מה נשמר
    ws_periods = wb['2️⃣ תקופות מילואים']
    ws_btl = wb['3️⃣ תשלומי ב"ל']
    
    periods_count = ws_periods.max_row - 1
    btl_count = ws_btl.max_row - 1
    
    print(f"\n✅ נשמרים:")
    print(f"   📋 תקופות מילואים: {periods_count} שורות")
    print(f"      (כולל כל העמודות: מזהה, תאריכים, ימים, תשלומים, הערות)")
    print(f"   💰 תשלומי ב\"ל: {btl_count} שורות")
    print(f"      (כל התשלומים לפני 28/11/2025)")
    
    # 2. מחיקת דוח מסכם
    ws_summary = wb['4️⃣ דוח מסכם']
    summary_rows = ws_summary.max_row - 1
    
    print(f"\n🗑️ נמחקים:")
    print(f"   📊 דוח מסכם: {summary_rows} שורות")
    
    for row in range(ws_summary.max_row, 1, -1):
        ws_summary.delete_rows(row)
    
    # 3. מחיקת רשימת תשלומים
    ws_payments = wb['💵 רשימת תשלומים']
    payments_rows = ws_payments.max_row - 1
    
    print(f"   💵 רשימת תשלומים: {payments_rows} מנות")
    
    for row in range(ws_payments.max_row, 1, -1):
        ws_payments.delete_rows(row)
    
    # שמירה
    wb.save(SYSTEM_FILE)
    
    print("\n" + "=" * 60)
    print("✅ ניקוי הושלם בהצלחה!")
    print("=" * 60)
    print("\n📝 מה הלאה:")
    print("   1. הפעל את מערכת המילואים")
    print("   2. לחץ על 'חישוב מלא'")
    print("   3. הדוח המסכם ורשימת התשלומים יחושבו מחדש")
    print(f"\n💾 גיבוי נשמר ב:\n   {backup_path}")
    
    input("\nלחץ Enter לסגירה...")

if __name__ == "__main__":
    try:
        clean_summary_and_payments()
    except Exception as e:
        print(f"\n❌ שגיאה: {e}")
        input("\nלחץ Enter לסגירה...")
