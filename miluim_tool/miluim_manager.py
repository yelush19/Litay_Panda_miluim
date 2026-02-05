#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
מערכת ניהול תשלומי מילואים מלאה - ליטאי ניהול שירותים
Complete System v2.0 with Color Coding
"""

import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timedelta
import os
import shutil
import calendar

SYSTEM_FILE = r"C:\Projects\LitayPandaMiluim\מערכת_מילואים_מלאה.xlsx"

# צבעי ליטאי
LITAY_GREEN = "#528163"
LITAY_GREEN_DARK = "#2d5f3f"
LITAY_GREEN_LIGHT = "#8dd1bb"
LITAY_BG = "#f5f6fa"

# צבעי סימון
COLOR_NEW = "D4EDDA"      # ירוק בהיר - שורה חדשה
COLOR_UPDATED = "FFF3CD"  # כתום בהיר - עודכן
COLOR_SKIPPED = "E2E3E5"  # אפור - דולג

# חגים יהודיים 2025
JEWISH_HOLIDAYS_2025 = [
    datetime(2025, 4, 13), datetime(2025, 4, 14), datetime(2025, 4, 19), datetime(2025, 4, 20),
    datetime(2025, 6, 2), datetime(2025, 6, 3),
    datetime(2025, 9, 23), datetime(2025, 9, 24),
    datetime(2025, 10, 2),
    datetime(2025, 10, 7), datetime(2025, 10, 8), datetime(2025, 10, 13), datetime(2025, 10, 14),
]

class MiluimManager:
    def __init__(self, root):
        self.root = root
        self.root.title("Miluim System - Litay")
        self.root.geometry("520x620")
        self.root.configure(bg=LITAY_BG)
        
        title = tk.Label(root, text="מערכת ניהול תשלומי מילואים",
                        font=("Arial", 18, "bold"), bg=LITAY_BG, fg=LITAY_GREEN_DARK)
        title.pack(pady=15)
        
        subtitle = tk.Label(root, text="Litay Management Services",
                           font=("Arial", 11), bg=LITAY_BG, fg=LITAY_GREEN)
        subtitle.pack()
        
        btn_frame = tk.Frame(root, bg=LITAY_BG)
        btn_frame.pack(pady=20, padx=40, fill="both", expand=True)
        
        self.create_button(btn_frame, "📥 Import MECANO / ייבוא מקאנו", self.import_mecano)
        self.create_button(btn_frame, "💰 Import BTL Payment / ייבוא תשלום ב״ל", self.import_btl)
        self.create_button(btn_frame, "➕ Import 40% Bonus / ייבוא תוספת 40%", self.import_40_percent)
        self.create_button(btn_frame, "🔄 Calculate All / חישוב מלא", self.calculate_all)
        self.create_button(btn_frame, "🔄 Sync BTL → Periods / סנכרון ב״ל לתקופות", self.sync_btl_to_periods)
        self.create_button(btn_frame, "📄 Unpaid Report / דוח הפרשים לתשלום", self.generate_unpaid_report)
        
        # כפתור איפוס באדום
        reset_btn = tk.Button(btn_frame, text="🗑️ Clear & Restart / מחיקה והתחלה מחדש", 
                              font=("Arial", 10, "bold"), bg="#e74c3c", fg="white",
                              activebackground="#c0392b", activeforeground="white",
                              cursor="hand2", command=self.clear_and_restart, height=2)
        reset_btn.pack(fill="x", pady=(15, 6))
        
        self.status_var = tk.StringVar(value="Ready / מוכן לעבודה")
        status = tk.Label(root, textvariable=self.status_var, font=("Arial", 10),
                         bg=LITAY_GREEN_LIGHT, fg=LITAY_GREEN_DARK, pady=10)
        status.pack(fill="x", side="bottom")
        
        self.update_all = None
        
    def create_button(self, parent, text, command):
        btn = tk.Button(parent, text=text, font=("Arial", 11), bg=LITAY_GREEN, fg="white",
                       activebackground=LITAY_GREEN_DARK, activeforeground="white",
                       cursor="hand2", command=command, height=2)
        btn.pack(fill="x", pady=6)
    
    def color_row(self, ws, row_num, color_code):
        """צביעת שורה שלמה"""
        fill = PatternFill(start_color=color_code, end_color=color_code, fill_type="solid")
        for col in range(1, ws.max_column + 1):
            ws.cell(row_num, col).fill = fill
    
    def get_tracking_sheet_name(self, wb):
        """זיהוי שם גיליון המעקב (תומך בשם ישן וחדש)"""
        if '📊 מעקב מילואים ותשלומים' in wb.sheetnames:
            return '📊 מעקב מילואים ותשלומים'
        elif '2️⃣ תקופות מילואים' in wb.sheetnames:
            return '2️⃣ תקופות מילואים'
        else:
            raise Exception("לא נמצא גיליון מעקב מילואים!")
    
    def normalize_name(self, name):
        if pd.isna(name):
            return ""
        return ' '.join(str(name).strip().split())
    
    def format_date(self, date_val):
        if pd.isna(date_val):
            return ""
        if isinstance(date_val, datetime):
            return date_val.strftime('%d/%m/%Y')
        date_str = str(date_val).strip()
        if len(date_str) == 10 and date_str[2] == '.' and date_str[5] == '.':
            parts = date_str.split('.')
            return f"{parts[0]}/{parts[1]}/{parts[2]}"
        if len(date_str) == 8 and date_str[2] == '/' and date_str[5] == '/':
            parts = date_str.split('/')
            year = '20' + parts[2] if int(parts[2]) < 50 else '19' + parts[2]
            return f"{parts[0]}/{parts[1]}/{year}"
        if hasattr(date_val, 'strftime'):
            return date_val.strftime('%d/%m/%Y')
        return date_str
    
    def normalize_date(self, date_val):
        formatted = self.format_date(date_val)
        return formatted.strip() if formatted else ""
    
    def parse_date(self, date_str):
        if pd.isna(date_str) or not date_str:
            return None
        if isinstance(date_str, datetime):
            return date_str
        try:
            parts = str(date_str).split('/')
            if len(parts) == 3:
                return datetime(int(parts[2]), int(parts[1]), int(parts[0]))
        except:
            pass
        return None
        
    def backup_file(self):
        if os.path.exists(SYSTEM_FILE):
            backup_dir = os.path.join(os.path.dirname(SYSTEM_FILE), "backups")
            os.makedirs(backup_dir, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = os.path.join(backup_dir, f"backup_{timestamp}.xlsx")
            shutil.copy2(SYSTEM_FILE, backup_path)
            return backup_path
        return None
    
    def count_work_days(self, start_date, end_date):
        if not start_date or not end_date:
            return 0, 0, 0, 0
        
        weekdays = 0
        fridays = 0
        saturdays = 0
        holidays = 0
        
        current = start_date
        while current <= end_date:
            weekday = current.weekday()
            if current in JEWISH_HOLIDAYS_2025:
                holidays += 1
            elif weekday == 5:
                saturdays += 1
            elif weekday == 4:
                fridays += 1
            else:
                weekdays += 1
            current += timedelta(days=1)
        
        return weekdays, fridays, saturdays, holidays
    
    def get_next_period_id(self, ws):
        max_id = 0
        for row in range(2, ws.max_row + 1):
            cell_val = ws.cell(row, 1).value
            if cell_val and str(cell_val).startswith('P'):
                try:
                    num = int(str(cell_val).replace('P', ''))
                    max_id = max(max_id, num)
                except:
                    pass
        return f"P{max_id + 1:04d}"
    
    def ask_name_mapping(self, mecano_name, system_names):
        dialog = tk.Toplevel(self.root)
        dialog.title("Name Mapping")
        dialog.geometry("400x320")
        dialog.configure(bg=LITAY_BG)
        dialog.transient(self.root)
        dialog.grab_set()
        
        result = {"choice": None}
        
        msg = f"Employee from MECANO:\n{mecano_name}\n\nNot found. Map to existing?"
        
        label = tk.Label(dialog, text=msg, font=("Arial", 10), bg=LITAY_BG, 
                        fg=LITAY_GREEN_DARK, justify="left")
        label.pack(pady=15, padx=15)
        
        listbox = tk.Listbox(dialog, font=("Arial", 10), height=7)
        listbox.pack(fill="both", expand=True, padx=15)
        
        for name in sorted(system_names):
            listbox.insert(tk.END, name)
        
        btn_frame = tk.Frame(dialog, bg=LITAY_BG)
        btn_frame.pack(pady=10)
        
        def on_select():
            selection = listbox.curselection()
            if selection:
                result["choice"] = listbox.get(selection[0])
            dialog.destroy()
        
        def on_new():
            result["choice"] = "NEW"
            dialog.destroy()
        
        def on_skip():
            result["choice"] = None
            dialog.destroy()
        
        tk.Button(btn_frame, text="Select", command=on_select, bg=LITAY_GREEN, fg="white", width=9).pack(side="left", padx=3)
        tk.Button(btn_frame, text="New", command=on_new, bg=LITAY_GREEN_DARK, fg="white", width=9).pack(side="left", padx=3)
        tk.Button(btn_frame, text="Skip", command=on_skip, bg="#999", fg="white", width=9).pack(side="left", padx=3)
        
        dialog.wait_window()
        return result["choice"]
    
    def split_period_by_month(self, start_date, end_date):
        """פיצול תקופה לפי חודשים קלנדריים"""
        periods = []
        current_start = start_date
        
        while current_start <= end_date:
            year = current_start.year
            month = current_start.month
            if month == 12:
                next_month_start = datetime(year + 1, 1, 1)
            else:
                next_month_start = datetime(year, month + 1, 1)
            
            month_end = next_month_start - timedelta(days=1)
            period_end = min(month_end, end_date)
            days = (period_end - current_start).days + 1
            
            periods.append({
                'start': current_start,
                'end': period_end,
                'days': days
            })
            
            current_start = period_end + timedelta(days=1)
        
        return periods
    
    def import_mecano(self):
        file_path = filedialog.askopenfilename(title="Select MECANO file",
                                               filetypes=[("Excel files", "*.xlsx *.xls")])
        if not file_path:
            return
        try:
            self.status_var.set("Importing MECANO...")
            self.root.update()
            
            df = pd.read_excel(file_path)
            df['תאריך'] = pd.to_datetime(df['תאריך'], format='%d.%m.%Y')
            df = df.sort_values(['שם עובד', 'תאריך'])
            
            # שלב 1: קיבוץ ימים רצופים
            raw_periods = []
            current_employee = None
            current_start = None
            current_end = None
            current_dept = None
            
            for _, row in df.iterrows():
                employee = self.normalize_name(row['שם עובד'])
                date = row['תאריך']
                dept = row['מחלקה']
                
                if employee != current_employee:
                    if current_employee:
                        raw_periods.append({
                            'עובד': current_employee,
                            'מחלקה': current_dept,
                            'התחלה': current_start,
                            'סיום': current_end
                        })
                    current_employee = employee
                    current_start = date
                    current_end = date
                    current_dept = dept
                else:
                    if (date - current_end).days == 1:
                        current_end = date
                    else:
                        raw_periods.append({
                            'עובד': current_employee,
                            'מחלקה': current_dept,
                            'התחלה': current_start,
                            'סיום': current_end
                        })
                        current_start = date
                        current_end = date
            
            if current_employee:
                raw_periods.append({
                    'עובד': current_employee,
                    'מחלקה': current_dept,
                    'התחלה': current_start,
                    'סיום': current_end
                })
            
            # שלב 2: פיצול לפי חודשים קלנדריים
            periods = []
            for raw_period in raw_periods:
                monthly_splits = self.split_period_by_month(
                    raw_period['התחלה'], 
                    raw_period['סיום']
                )
                
                for split in monthly_splits:
                    periods.append({
                        'עובד': raw_period['עובד'],
                        'מחלקה': raw_period['מחלקה'],
                        'התחלה': split['start'],
                        'סיום': split['end'],
                        'ימים': split['days']
                    })
            
            self.backup_file()
            
            wb = load_workbook(SYSTEM_FILE)
            tracking_sheet = self.get_tracking_sheet_name(wb)
            ws_periods = wb[tracking_sheet]
            ws_employees = wb['1️⃣ רשימת עובדים']
            
            df_employees = pd.read_excel(SYSTEM_FILE, sheet_name='1️⃣ רשימת עובדים')
            system_names = set(df_employees['שם מלא'].dropna().apply(self.normalize_name))
            employee_rates = dict(zip(df_employees['שם מלא'].apply(self.normalize_name), 
                                     df_employees['תעריף יומי']))
            
            existing_periods = {}
            for row in range(2, ws_periods.max_row + 1):
                emp = self.normalize_name(ws_periods.cell(row, 2).value)
                start = self.normalize_date(ws_periods.cell(row, 4).value)
                end = self.normalize_date(ws_periods.cell(row, 5).value)
                key = f"{emp}|{start}|{end}"
                existing_periods[key] = row
            
            added = 0
            skipped = 0
            new_employees = []
            name_mappings = {}
            
            next_row = ws_periods.max_row + 1
            
            for period in periods:
                emp_name = period['עובד']
                
                if emp_name not in system_names and emp_name not in name_mappings:
                    choice = self.ask_name_mapping(emp_name, system_names)
                    if choice == "NEW":
                        new_employees.append(emp_name)
                        system_names.add(emp_name)
                        name_mappings[emp_name] = emp_name
                    elif choice:
                        name_mappings[emp_name] = choice
                    else:
                        skipped += 1
                        continue
                
                final_name = name_mappings.get(emp_name, emp_name)
                
                start_str = self.format_date(period['התחלה'])
                end_str = self.format_date(period['סיום'])
                key = f"{final_name}|{start_str}|{end_str}"
                
                if key in existing_periods:
                    skipped += 1
                    continue
                
                weekdays, fridays, saturdays, holidays = self.count_work_days(
                    period['התחלה'], period['סיום'])
                
                period_id = self.get_next_period_id(ws_periods)
                ws_periods.cell(next_row, 1).value = period_id
                ws_periods.cell(next_row, 2).value = final_name
                ws_periods.cell(next_row, 3).value = period['מחלקה']
                ws_periods.cell(next_row, 4).value = start_str
                ws_periods.cell(next_row, 5).value = end_str
                ws_periods.cell(next_row, 6).value = period['התחלה'].strftime('%m/%Y')
                ws_periods.cell(next_row, 7).value = period['ימים']
                ws_periods.cell(next_row, 8).value = weekdays
                ws_periods.cell(next_row, 9).value = fridays
                ws_periods.cell(next_row, 10).value = saturdays
                ws_periods.cell(next_row, 11).value = holidays
                
                rate = employee_rates.get(final_name, 0)
                ws_periods.cell(next_row, 12).value = rate
                
                if weekdays > 0:
                    ws_periods.cell(next_row, 13).value = weekdays * rate
                
                # צביעה בירוק - שורה חדשה
                self.color_row(ws_periods, next_row, COLOR_NEW)
                
                next_row += 1
                added += 1
            
            if new_employees:
                next_emp_row = ws_employees.max_row + 1
                for emp_name in new_employees:
                    ws_employees.cell(next_emp_row, 4).value = emp_name
                    ws_employees.cell(next_emp_row, 10).value = "פעיל"
                    # צביעה בירוק
                    self.color_row(ws_employees, next_emp_row, COLOR_NEW)
                    next_emp_row += 1
            
            wb.save(SYSTEM_FILE)
            
            self.status_var.set(f"MECANO: {added} added, {skipped} skipped")
            messagebox.showinfo("Success", 
                f"MECANO Import Complete\n\n"
                f"Records: {len(df)}\n"
                f"Periods: {len(periods)}\n\n"
                f"✅ Added: {added} (green)\n"
                f"⏭️ Skipped: {skipped}\n"
                f"👤 New employees: {len(new_employees)}")
            
        except Exception as e:
            self.status_var.set("Error")
            messagebox.showerror("Error", f"MECANO Error:\n{str(e)}")
    
    def get_existing_btl_records(self, ws):
        existing = {}
        for row in range(2, ws.max_row + 1):
            emp = self.normalize_name(ws.cell(row, 2).value)
            start_date = self.normalize_date(ws.cell(row, 3).value)
            end_date = self.normalize_date(ws.cell(row, 4).value)
            claim_type = str(ws.cell(row, 5).value or "").strip()
            tagmul = ws.cell(row, 6).value or 0
            if emp:
                key = f"{emp}|{start_date}|{end_date}|{claim_type}"
                existing[key] = {"row": row, "tagmul": tagmul}
        return existing
    
    def ask_update_or_skip(self, employee_name, date_start, existing_amount, new_amount):
        if self.update_all is not None:
            return self.update_all
        
        dialog = tk.Toplevel(self.root)
        dialog.title("Duplicate")
        dialog.geometry("420x230")
        dialog.configure(bg=LITAY_BG)
        dialog.transient(self.root)
        dialog.grab_set()
        
        result = {"choice": "skip"}
        
        msg = f"""Found existing with different amount:

{employee_name} | {date_start}
Existing: {existing_amount:,.0f} NIS
New: {new_amount:,.0f} NIS"""
        
        label = tk.Label(dialog, text=msg, font=("Arial", 10), bg=LITAY_BG, 
                        fg=LITAY_GREEN_DARK, justify="left")
        label.pack(pady=15, padx=15)
        
        btn_frame = tk.Frame(dialog, bg=LITAY_BG)
        btn_frame.pack(pady=10)
        
        def on_update():
            result["choice"] = "update"
            dialog.destroy()
        def on_skip():
            result["choice"] = "skip"
            dialog.destroy()
        def on_update_all():
            result["choice"] = "update"
            self.update_all = "update"
            dialog.destroy()
        def on_skip_all():
            result["choice"] = "skip"
            self.update_all = "skip"
            dialog.destroy()
        
        tk.Button(btn_frame, text="Update", font=("Arial", 9), bg=LITAY_GREEN, fg="white",
                 command=on_update, width=9).grid(row=0, column=0, padx=4, pady=4)
        tk.Button(btn_frame, text="Skip", font=("Arial", 9), bg="#999", fg="white",
                 command=on_skip, width=9).grid(row=0, column=1, padx=4, pady=4)
        tk.Button(btn_frame, text="Update All", font=("Arial", 9), bg=LITAY_GREEN_DARK, fg="white",
                 command=on_update_all, width=9).grid(row=1, column=0, padx=4, pady=4)
        tk.Button(btn_frame, text="Skip All", font=("Arial", 9), bg="#666", fg="white",
                 command=on_skip_all, width=9).grid(row=1, column=1, padx=4, pady=4)
        
        dialog.wait_window()
        return result["choice"]
            
    def import_btl(self):
        file_path = filedialog.askopenfilename(title="Select BTL file",
                                               filetypes=[("Excel files", "*.xlsx *.xls *.xla")])
        if not file_path:
            return
        try:
            self.status_var.set("Importing BTL...")
            self.root.update()
            
            self.update_all = None
            
            df = pd.read_excel(file_path, header=None)
            
            mana_number = df.iloc[2, 1]
            payment_date = df.iloc[9, 1]
            
            headers = df.iloc[11].tolist()
            data = df.iloc[12:].copy()
            data.columns = headers
            data = data.dropna(subset=['זהות'])
            
            self.backup_file()
            
            wb = load_workbook(SYSTEM_FILE)
            ws = wb['3️⃣ תשלומי ב"ל']
            ws_payments = wb['💵 רשימת תשלומים']
            
            existing = self.get_existing_btl_records(ws)
            
            next_row = ws.max_row + 1
            
            added = 0
            skipped = 0
            updated = 0
            total_tagmul = 0
            total_pitzuy = 0
            
            for _, row in data.iterrows():
                try:
                    tz = str(row['זהות']).strip()
                    employee_name = self.normalize_name(f"{row['שם פרטי']} {row['שם משפחה']}")
                    start_date = self.normalize_date(row['תאריך שרות'])
                    end_date = self.normalize_date(row['תאריך סיום שרות'])
                    claim_type = str(row['סוג תביעה']).strip()
                    
                    tagmul_raw = row['תגמול']
                    pitzuy_raw = row['פיצוי %20 למעסיק']
                    
                    tagmul = 0
                    if pd.notna(tagmul_raw):
                        tagmul_str = str(tagmul_raw).replace('+', '').replace('-', '').replace(',', '')
                        if tagmul_str and not str(tagmul_raw).startswith('-'):
                            tagmul = float(tagmul_str) if tagmul_str else 0
                    
                    pitzuy = 0
                    if pd.notna(pitzuy_raw):
                        pitzuy_str = str(pitzuy_raw).replace('+', '').replace('-', '').replace(',', '')
                        if pitzuy_str and not str(pitzuy_raw).startswith('-'):
                            pitzuy = float(pitzuy_str) if pitzuy_str else 0
                    
                    key = f"{employee_name}|{start_date}|{end_date}|{claim_type}"
                    
                    if key in existing:
                        existing_tagmul = existing[key]["tagmul"] or 0
                        existing_row = existing[key]["row"]
                        
                        if abs(existing_tagmul - tagmul) < 1:
                            skipped += 1
                            continue
                        else:
                            choice = self.ask_update_or_skip(employee_name, start_date, 
                                                            existing_tagmul, tagmul)
                            if choice == "skip":
                                skipped += 1
                                continue
                            else:
                                ws.cell(existing_row, 6).value = tagmul
                                ws.cell(existing_row, 7).value = pitzuy
                                ws.cell(existing_row, 9).value = tagmul
                                ws.cell(existing_row, 10).value = mana_number
                                ws.cell(existing_row, 11).value = self.format_date(payment_date)
                                # צביעה בכתום - עודכן
                                self.color_row(ws, existing_row, COLOR_UPDATED)
                                updated += 1
                                total_tagmul += tagmul
                                total_pitzuy += pitzuy
                                continue
                    
                    ws.cell(next_row, 1).value = tz
                    ws.cell(next_row, 2).value = employee_name
                    ws.cell(next_row, 3).value = start_date
                    ws.cell(next_row, 4).value = end_date
                    ws.cell(next_row, 5).value = claim_type
                    ws.cell(next_row, 6).value = tagmul
                    ws.cell(next_row, 7).value = pitzuy
                    ws.cell(next_row, 8).value = 0
                    ws.cell(next_row, 9).value = tagmul
                    ws.cell(next_row, 10).value = mana_number
                    ws.cell(next_row, 11).value = self.format_date(payment_date)
                    ws.cell(next_row, 12).value = os.path.basename(file_path)
                    
                    # צביעה בירוק - שורה חדשה
                    self.color_row(ws, next_row, COLOR_NEW)
                    
                    total_tagmul += tagmul
                    total_pitzuy += pitzuy
                    next_row += 1
                    added += 1
                    
                except Exception as e:
                    print(f"Row error: {e}")
                    continue
            
            # עדכון רשימת תשלומים
            mana_exists = False
            for r in range(2, ws_payments.max_row + 1):
                if ws_payments.cell(r, 1).value == mana_number:
                    mana_exists = True
                    ws_payments.cell(r, 3).value = total_tagmul
                    ws_payments.cell(r, 4).value = total_pitzuy
                    ws_payments.cell(r, 6).value = total_tagmul + total_pitzuy
                    # צביעה בכתום
                    self.color_row(ws_payments, r, COLOR_UPDATED)
                    break
            
            if not mana_exists:
                next_payment_row = ws_payments.max_row + 1
                ws_payments.cell(next_payment_row, 1).value = mana_number
                ws_payments.cell(next_payment_row, 2).value = self.format_date(payment_date)
                ws_payments.cell(next_payment_row, 3).value = total_tagmul
                ws_payments.cell(next_payment_row, 4).value = total_pitzuy
                ws_payments.cell(next_payment_row, 5).value = 0
                ws_payments.cell(next_payment_row, 6).value = total_tagmul + total_pitzuy
                # צביעה בירוק
                self.color_row(ws_payments, next_payment_row, COLOR_NEW)
            
            wb.save(SYSTEM_FILE)
            
            self.status_var.set(f"BTL: {added} added, {updated} updated")
            messagebox.showinfo("Success", 
                f"Mana: {mana_number} | {self.format_date(payment_date)}\n\n"
                f"✅ Added: {added} (green)\n"
                f"🔄 Updated: {updated} (orange)\n"
                f"⏭️ Skipped: {skipped}\n\n"
                f"Tagmul: {total_tagmul:,.0f}\n"
                f"Pitzuy: {total_pitzuy:,.0f}\n"
                f"Total: {total_tagmul + total_pitzuy:,.0f} NIS")
            
        except Exception as e:
            self.status_var.set("Error")
            messagebox.showerror("Error", f"BTL Error:\n{str(e)}")
            
    def import_40_percent(self):
        file_path = filedialog.askopenfilename(title="Select 40% Bonus file",
                                               filetypes=[("Excel files", "*.xlsx *.xls *.xla")])
        if not file_path:
            return
        try:
            self.status_var.set("Importing 40%...")
            self.root.update()
            
            self.update_all = None
            
            df = pd.read_excel(file_path, header=None)
            
            mana_number = df.iloc[2, 1]
            payment_date = df.iloc[9, 1]
            
            headers = df.iloc[11].tolist()
            data = df.iloc[12:].copy()
            data.columns = headers
            data = data.dropna(subset=['זהות'])
            
            self.backup_file()
            
            wb = load_workbook(SYSTEM_FILE)
            ws = wb['3️⃣ תשלומי ב"ל']
            ws_payments = wb['💵 רשימת תשלומים']
            
            existing = self.get_existing_btl_records(ws)
            
            next_row = ws.max_row + 1
            
            added = 0
            skipped = 0
            total_40 = 0
            
            for _, row in data.iterrows():
                try:
                    tz = str(row['זהות']).strip()
                    employee_name = self.normalize_name(f"{row['שם פרטי']} {row['שם משפחה']}")
                    start_date = self.normalize_date(row['תאריך שרות'])
                    end_date = self.normalize_date(row['תאריך סיום שרות'])
                    
                    claim_type = "תוספת 40%"
                    
                    bonus_40_raw = row.get('תגמול נדרש', row.get('תגמול', 0))
                    
                    bonus_40 = 0
                    if pd.notna(bonus_40_raw):
                        bonus_str = str(bonus_40_raw).replace('+', '').replace('-', '').replace(',', '')
                        if bonus_str and not str(bonus_40_raw).startswith('-'):
                            bonus_40 = float(bonus_str) if bonus_str else 0
                    
                    if bonus_40 == 0:
                        continue
                    
                    key = f"{employee_name}|{start_date}|{end_date}|{claim_type}"
                    
                    if key in existing:
                        skipped += 1
                        continue
                    
                    ws.cell(next_row, 1).value = tz
                    ws.cell(next_row, 2).value = employee_name
                    ws.cell(next_row, 3).value = start_date
                    ws.cell(next_row, 4).value = end_date
                    ws.cell(next_row, 5).value = claim_type
                    ws.cell(next_row, 6).value = 0
                    ws.cell(next_row, 7).value = 0
                    ws.cell(next_row, 8).value = bonus_40
                    ws.cell(next_row, 9).value = bonus_40
                    ws.cell(next_row, 10).value = mana_number
                    ws.cell(next_row, 11).value = self.format_date(payment_date)
                    ws.cell(next_row, 12).value = os.path.basename(file_path)
                    
                    # צביעה בירוק
                    self.color_row(ws, next_row, COLOR_NEW)
                    
                    total_40 += bonus_40
                    next_row += 1
                    added += 1
                    
                except Exception as e:
                    print(f"Row error: {e}")
                    continue
            
            for r in range(2, ws_payments.max_row + 1):
                if ws_payments.cell(r, 1).value == mana_number:
                    ws_payments.cell(r, 5).value = total_40
                    current_total = (ws_payments.cell(r, 3).value or 0) + \
                                   (ws_payments.cell(r, 4).value or 0) + total_40
                    ws_payments.cell(r, 6).value = current_total
                    self.color_row(ws_payments, r, COLOR_UPDATED)
                    break
            
            wb.save(SYSTEM_FILE)
            
            self.status_var.set(f"40%: {added} added")
            messagebox.showinfo("Success", 
                f"40% Bonus Import\n\n"
                f"Mana: {mana_number}\n"
                f"✅ Added: {added} (green)\n"
                f"Total 40%: {total_40:,.0f} NIS")
            
        except Exception as e:
            self.status_var.set("Error")
            messagebox.showerror("Error", f"40% Error:\n{str(e)}")
        
    def calculate_all(self):
        """חישוב מלא - לפי תקופות בודדות"""
        try:
            self.status_var.set("Calculating...")
            self.root.update()
            
            self.backup_file()
            
            wb = load_workbook(SYSTEM_FILE)
            tracking_sheet = self.get_tracking_sheet_name(wb)
            ws_periods = wb[tracking_sheet]
            ws_btl = wb['3️⃣ תשלומי ב"ל']
            ws_summary = wb['4️⃣ דוח מסכם']
            ws_employees = wb['1️⃣ רשימת עובדים']
            
            # קריאת תעריפים
            df_employees = pd.read_excel(SYSTEM_FILE, sheet_name='1️⃣ רשימת עובדים')
            employee_data = {}
            for _, emp in df_employees.iterrows():
                name = self.normalize_name(emp['שם מלא'])
                employee_data[name] = {
                    'rate': emp.get('תעריף יומי', 0),
                    'monthly': emp.get('משכורת חודשית', 0)
                }
            
            # קריאת תקופות - כל תקופה בנפרד
            df_periods = pd.read_excel(SYSTEM_FILE, sheet_name=tracking_sheet)
            df_btl = pd.read_excel(SYSTEM_FILE, sheet_name='3️⃣ תשלומי ב"ל')
            
            summary_data = []
            
            # לולאה על כל תקופה (לא קיבוץ!)
            for _, period in df_periods.iterrows():
                emp = self.normalize_name(period['שם עובד'])
                period_id = period['מזהה תקופה']
                department = period.get('מחלקה', '')  # משיכת מחלקה
                start_date = period['תאריך התחלה']
                end_date = period['תאריך סיום']
                month = period.get('חודש', '')
                total_days = period['סה"כ ימים']
                weekdays = period['ימי א-ה']
                
                emp_info = employee_data.get(emp, {})
                rate = emp_info.get('rate', 0)
                monthly = emp_info.get('monthly', 0)
                
                # חישוב תשלום מעסיק
                if weekdays > 20:
                    employer_payment = monthly
                else:
                    employer_payment = weekdays * rate
                
                # משיכת תשלומי ב"ל - התאמה לפי תאריכים
                start_parsed = self.parse_date(start_date)
                end_parsed = self.parse_date(end_date)
                
                btl_payments = df_btl[
                    (df_btl['שם עובד'].apply(self.normalize_name) == emp) &
                    (df_btl['תאריך התחלה'].apply(self.parse_date) == start_parsed) &
                    (df_btl['תאריך סיום'].apply(self.parse_date) == end_parsed)
                ]
                
                btl_tagmul = btl_payments['תגמול ₪'].sum()
                btl_pitzuy = btl_payments['פיצוי 20% ₪'].sum()
                btl_40 = btl_payments['תוספת 40% ₪'].sum()
                
                # הפרש = תגמול ב"ל - תשלום מעסיק
                # חיובי = לטובת העובד (ב"ל שילם יותר)
                # שלילי = המעסיק שילם יותר
                difference = btl_tagmul - employer_payment
                
                summary_data.append({
                    'מזהה': period_id,
                    'עובד': emp,
                    'מחלקה': department,  # הוספת מחלקה
                    'חודש': month,
                    'התחלה': start_date,
                    'סיום': end_date,
                    'ימים': total_days,
                    'ימי א-ה': weekdays,
                    'תעריף': rate,
                    'תשלום מעסיק': employer_payment,
                    'תגמול ב"ל': btl_tagmul,
                    'פיצוי 20%': btl_pitzuy,
                    'תוספת 40%': btl_40,
                    'הפרש': difference
                })
            
            # ניקוי דוח מסכם
            for row in range(ws_summary.max_row, 1, -1):
                if row > 1:
                    ws_summary.delete_rows(row)
            
            next_row = 2
            for item in summary_data:
                ws_summary.cell(next_row, 1).value = item['עובד']
                ws_summary.cell(next_row, 2).value = item['מזהה']
                ws_summary.cell(next_row, 3).value = item['מחלקה']  # מחלקה
                ws_summary.cell(next_row, 4).value = item['חודש']
                ws_summary.cell(next_row, 5).value = item['התחלה']
                ws_summary.cell(next_row, 6).value = item['סיום']
                ws_summary.cell(next_row, 7).value = item['ימים']
                ws_summary.cell(next_row, 8).value = item['ימי א-ה']
                ws_summary.cell(next_row, 9).value = item['תעריף']
                ws_summary.cell(next_row, 10).value = item['תשלום מעסיק']
                ws_summary.cell(next_row, 11).value = item['תגמול ב"ל']
                ws_summary.cell(next_row, 12).value = item['פיצוי 20%']
                ws_summary.cell(next_row, 13).value = item['תוספת 40%']
                ws_summary.cell(next_row, 14).value = item['הפרש']
                
                # סטטוס לפי הפרש:
                # הפרש = 0 → מאוזן
                # הפרש > 0 → ב"ל שילם יותר → ממתין (צריך לשלם לעובד)
                # הפרש < 0 → מעסיק שילם יותר → לא רלוונטי
                if abs(item['הפרש']) < 1:
                    status = "מאוזן"
                elif item['הפרש'] > 0:
                    status = "ממתין"
                else:
                    status = "לא רלוונטי"
                ws_summary.cell(next_row, 15).value = status  # עמודה 15
                
                self.color_row(ws_summary, next_row, COLOR_NEW)
                
                next_row += 1
            
            wb.save(SYSTEM_FILE)
            
            total_employer = sum(x['תשלום מעסיק'] for x in summary_data)
            total_btl = sum(x['תגמול ב"ל'] for x in summary_data)
            total_diff = sum(x['הפרש'] for x in summary_data)
            
            self.status_var.set("Calculation complete")
            messagebox.showinfo("Success", 
                f"Calculation Complete\n\n"
                f"✅ Periods: {len(summary_data)} (green)\n\n"
                f"Employer: {total_employer:,.0f} NIS\n"
                f"BTL: {total_btl:,.0f} NIS\n"
                f"Difference: {total_diff:,.0f} NIS")
            
        except Exception as e:
            self.status_var.set("Error")
            messagebox.showerror("Error", f"Calculation Error:\n{str(e)}")
    
    def sync_btl_to_periods(self):
        """סנכרון נתוני ב"ל לטאב תקופות מילואים"""
        try:
            self.status_var.set("Syncing BTL data...")
            self.root.update()
            
            if not os.path.exists(SYSTEM_FILE):
                messagebox.showerror("Error", "System file not found!")
                return
            
            self.backup_file()
            
            # טעינה לכתיבה
            wb = load_workbook(SYSTEM_FILE)
            
            # טעינה לקריאת ערכים (לא נוסחאות)
            wb_read = load_workbook(SYSTEM_FILE, data_only=True)
            
            # זיהוי שם גיליון המעקב
            tracking_sheet = self.get_tracking_sheet_name(wb)
            ws_periods = wb[tracking_sheet]
            ws_periods_read = wb_read[tracking_sheet]
            ws_btl = wb['3️⃣ תשלומי ב"ל']
            
            # קריאת נתונים
            df_periods = pd.read_excel(SYSTEM_FILE, sheet_name=tracking_sheet)
            df_btl = pd.read_excel(SYSTEM_FILE, sheet_name='3️⃣ תשלומי ב"ל')
            
            updated_count = 0
            not_found_count = 0
            periods_without_btl = []
            btl_without_periods = []
            
            print("\n" + "=" * 60)
            print("🔄 סנכרון נתוני ב\"ל לתקופות מילואים")
            print("=" * 60)
            
            # שלב 1: מעבר על כל תקופה וחיפוש ב"ל תואם
            print("\n📊 שלב 1: עדכון תקופות מילואים...")
            for idx, period in df_periods.iterrows():
                period_id = period['מזהה תקופה']
                emp = self.normalize_name(period['שם עובד'])
                start_date = period['תאריך התחלה']
                end_date = period['תאריך סיום']
                
                # התאמה לפי שם + תאריכים
                start_parsed = self.parse_date(start_date)
                end_parsed = self.parse_date(end_date)
                
                btl_payments = df_btl[
                    (df_btl['שם עובד'].apply(self.normalize_name) == emp) &
                    (df_btl['תאריך התחלה'].apply(self.parse_date) == start_parsed) &
                    (df_btl['תאריך סיום'].apply(self.parse_date) == end_parsed)
                ]
                
                if len(btl_payments) > 0:
                    # עדכון השורה בטאב תקופות (idx+2 כי שורה 1 = כותרת)
                    row = idx + 2
                    
                    # סיכום כל התשלומים לתקופה זו
                    pitzuy = btl_payments['פיצוי 20% ₪'].sum()
                    tagmul = btl_payments['תגמול ₪'].sum()
                    bonus_40 = btl_payments['תוספת 40% ₪'].sum()
                    
                    # מועד תשלום - האחרון
                    payment_dates = btl_payments['תאריך תשלום'].dropna()
                    if len(payment_dates) > 0:
                        last_payment = payment_dates.iloc[-1]
                    else:
                        last_payment = None
                    
                    # עדכון עמודות
                    ws_periods.cell(row, 14).value = pitzuy  # פיצוי 20%
                    ws_periods.cell(row, 15).value = bonus_40  # תוספת 40%
                    ws_periods.cell(row, 16).value = tagmul  # סה"כ תגמול
                    ws_periods.cell(row, 17).value = last_payment  # מועד תשלום
                    
                    # חישוב הפרשים - קריאה מהגיליון עם ערכים מחושבים
                    employer_payment_raw = ws_periods_read.cell(row, 13).value
                    
                    # המרה למספר (טיפול בטקסט/None)
                    if employer_payment_raw is None or str(employer_payment_raw).strip() == '':
                        employer_payment = 0
                    else:
                        try:
                            # ניסיון להמיר למספר
                            employer_payment = float(str(employer_payment_raw).replace(',', ''))
                        except:
                            employer_payment = 0
                    
                    diff = tagmul - employer_payment
                    ws_periods.cell(row, 18).value = diff  # הפרש
                    
                    # צביעה בכתום
                    for col in range(14, 19):
                        ws_periods.cell(row, col).fill = PatternFill(
                            start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"
                        )
                    
                    updated_count += 1
                    if updated_count <= 5:  # הצג רק 5 ראשונים
                        print(f"   ✅ {period_id} | {emp[:20]:20} | תגמול: {tagmul:,.0f} ₪")
                else:
                    not_found_count += 1
                    periods_without_btl.append({
                        'מזהה': period_id,
                        'עובד': emp,
                        'התחלה': start_date,
                        'סיום': end_date
                    })
                    if not_found_count <= 3:  # הצג רק 3 ראשונים
                        print(f"   ⚠️  {period_id} | {emp[:20]:20} | אין תשלום ב\"ל")
            
            # שלב 2: חיפוש תשלומי ב"ל ללא תקופה תואמת
            print("\n🔍 שלב 2: בדיקת תשלומי ב\"ל ללא תקופה...")
            for idx, btl in df_btl.iterrows():
                emp = self.normalize_name(btl['שם עובד'])
                start_date = btl['תאריך התחלה']
                end_date = btl['תאריך סיום']
                tagmul = btl.get('תגמול ₪', 0)
                
                start_parsed = self.parse_date(start_date)
                end_parsed = self.parse_date(end_date)
                
                # חיפוש תקופה תואמת
                matching_periods = df_periods[
                    (df_periods['שם עובד'].apply(self.normalize_name) == emp) &
                    (df_periods['תאריך התחלה'].apply(self.parse_date) == start_parsed) &
                    (df_periods['תאריך סיום'].apply(self.parse_date) == end_parsed)
                ]
                
                if len(matching_periods) == 0:
                    # שורה יתומה - צביעה באדום!
                    btl_row = idx + 2  # שורה בטאב ב"ל
                    
                    # צביעה אדומה בהירה
                    red_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                    for col in range(1, ws_btl.max_column + 1):
                        ws_btl.cell(btl_row, col).fill = red_fill
                    
                    btl_without_periods.append({
                        'עובד': emp,
                        'התחלה': start_date,
                        'סיום': end_date,
                        'תגמול': tagmul,
                        'שורה': btl_row
                    })
                    
                    if len(btl_without_periods) <= 3:
                        print(f"   🔴 שורה {btl_row} | {emp[:20]:20} | {start_date} - {end_date}")
            
            if len(btl_without_periods) > 3:
                print(f"   ... ועוד {len(btl_without_periods) - 3} שורות יתומות")
            
            wb.save(SYSTEM_FILE)
            wb.close()
            wb_read.close()
            
            print("\n" + "=" * 60)
            print(f"✅ סנכרון הושלם!")
            print("=" * 60)
            print(f"\n📊 סיכום:")
            print(f"   ✅ תקופות שעודכנו: {updated_count}")
            print(f"   ⚠️  תקופות ללא ב\"ל: {not_found_count}")
            print(f"   🔍 תשלומי ב\"ל ללא תקופה: {len(btl_without_periods)}")
            
            # הצגת דוח מפורט
            message = f"BTL Sync Complete!\n\n"
            message += f"✅ Updated: {updated_count} periods\n"
            message += f"⚠️ Periods without BTL: {not_found_count}\n"
            message += f"🔍 BTL without periods: {len(btl_without_periods)}\n\n"
            
            if len(btl_without_periods) > 0:
                message += f"⚠️ Found {len(btl_without_periods)} BTL payments without matching periods!\n\n"
                message += "🔴 These rows are marked in RED in the BTL tab.\n\n"
                message += "These payments are in BTL tab but not in Periods tab.\n"
                message += "This might indicate:\n"
                message += "• Missing MECANO import\n"
                message += "• Name/date mismatch\n"
                message += "• Duplicate BTL payments\n\n"
                
                # הצג 3 ראשונים
                message += "First 3 examples:\n"
                for i, item in enumerate(btl_without_periods[:3], 1):
                    message += f"🔴 Row {item['שורה']} | {item['עובד'][:25]} | {item['התחלה']} - {item['סיום']} | {item['תגמול']:,.0f} ₪\n"
                
                if len(btl_without_periods) > 3:
                    message += f"... and {len(btl_without_periods) - 3} more\n"
            
            self.status_var.set(f"Synced: {updated_count} | Orphan BTL: {len(btl_without_periods)}")
            
            # הצגת חלון עם התוצאות
            result_window = tk.Toplevel(self.root)
            result_window.title("🔄 Sync Results")
            result_window.geometry("700x500")
            result_window.configure(bg=LITAY_BG)
            
            # טקסט עם תוצאות
            text_frame = tk.Frame(result_window, bg=LITAY_BG)
            text_frame.pack(pady=10, padx=10, fill="both", expand=True)
            
            text = tk.Text(text_frame, wrap="word", font=("Arial", 10), bg="white")
            text.pack(side="left", fill="both", expand=True)
            
            scrollbar = tk.Scrollbar(text_frame, command=text.yview)
            scrollbar.pack(side="right", fill="y")
            text.config(yscrollcommand=scrollbar.set)
            
            # כתיבת התוצאות
            text.insert("1.0", message)
            
            if len(btl_without_periods) > 3:
                text.insert("end", "\n" + "=" * 60 + "\n")
                text.insert("end", "🔴 Full list of BTL payments without periods (marked in RED):\n")
                text.insert("end", "=" * 60 + "\n\n")
                for i, item in enumerate(btl_without_periods, 1):
                    text.insert("end", f"🔴 Row {item['שורה']:3} | {item['עובד']:30} | {item['התחלה']} - {item['סיום']} | {item['תגמול']:,.0f} ₪\n")
            
            text.config(state="disabled")
            
            # כפתור סגירה
            close_btn = tk.Button(result_window, text="Close", command=result_window.destroy,
                                  bg=LITAY_GREEN, fg="white", font=("Arial", 11, "bold"), 
                                  width=15, height=2)
            close_btn.pack(pady=10)
            
        except Exception as e:
            self.status_var.set("Error")
            messagebox.showerror("Error", f"Sync Error:\n{str(e)}")
            import traceback
            traceback.print_exc()
    
    def generate_unpaid_report(self):
        """הפקת דוח הפרשים שטרם שולמו"""
        try:
            self.status_var.set("Generating unpaid report...")
            self.root.update()
            
            if not os.path.exists(SYSTEM_FILE):
                messagebox.showerror("Error", "System file not found!")
                return
            
            # בדיקה איזה גיליון קיים
            wb = load_workbook(SYSTEM_FILE)
            sheet_name = self.get_tracking_sheet_name(wb)
            ws = wb[sheet_name]
            
            # ספירת שורות ללא חודש ביצוע תשלום
            unpaid_rows = []
            for row in range(2, ws.max_row + 1):
                period_id = ws.cell(row, 1).value
                payment_month = ws.cell(row, 20).value  # עמודה T - חודש ביצוע תשלום
                
                if period_id and (not payment_month or str(payment_month).strip() == ''):
                    unpaid_rows.append(row)
            
            wb.close()
            
            if len(unpaid_rows) == 0:
                self.status_var.set("No unpaid items")
                messagebox.showinfo("Info", 
                    "No unpaid differences found!\n\n"
                    "All periods have payment month assigned.")
                return
            
            # יצירת קובץ חדש
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_dir = os.path.dirname(SYSTEM_FILE)
            output_file = os.path.join(output_dir, f"דוח_הפרשים_לתשלום_{timestamp}.xlsx")
            
            # פתיחה מחדש לעריכה
            wb = load_workbook(SYSTEM_FILE)
            
            # מחיקת גיליונות מיותרים
            sheets_to_keep = [sheet_name]
            for sheet in wb.sheetnames:
                if sheet not in sheets_to_keep:
                    del wb[sheet]
            
            ws = wb[sheet_name]
            
            # מחיקת שורות ששולמו (מלמטה למעלה)
            all_rows = list(range(2, ws.max_row + 1))
            paid_rows = [r for r in all_rows if r not in unpaid_rows]
            
            for row in reversed(paid_rows):
                ws.delete_rows(row)
            
            wb.save(output_file)
            wb.close()
            
            self.status_var.set(f"Unpaid report: {len(unpaid_rows)} items")
            messagebox.showinfo("Success", 
                f"Unpaid Differences Report Created!\n\n"
                f"📄 Items: {len(unpaid_rows)}\n\n"
                f"File:\n{os.path.basename(output_file)}")
            
        except Exception as e:
            self.status_var.set("Error")
            messagebox.showerror("Error", f"Report Error:\n{str(e)}")
    
    def clear_and_restart(self):
        """מחיקת כל הנתונים והתחלה מחדש"""
        # חלון אישור
        dialog = tk.Toplevel(self.root)
        dialog.title("⚠️ Warning / אזהרה")
        dialog.geometry("450x280")
        dialog.configure(bg=LITAY_BG)
        dialog.transient(self.root)
        dialog.grab_set()
        
        result = {"choice": False}
        
        msg = """⚠️ WARNING / אזהרה ⚠️

This will DELETE ALL DATA from:
• Reserve periods (תקופות מילואים)
• BTL payments (תשלומי ב"ל)
• Payment list (רשימת תשלומים)
• Summary report (דוח מסכם)

Employee list will NOT be deleted.

A backup will be created automatically.

Are you sure?"""
        
        label = tk.Label(dialog, text=msg, font=("Arial", 10), bg=LITAY_BG, 
                        fg="#e74c3c", justify="left")
        label.pack(pady=20, padx=20)
        
        btn_frame = tk.Frame(dialog, bg=LITAY_BG)
        btn_frame.pack(pady=10)
        
        def on_confirm():
            result["choice"] = True
            dialog.destroy()
        
        def on_cancel():
            result["choice"] = False
            dialog.destroy()
        
        tk.Button(btn_frame, text="✅ Yes, Delete All", command=on_confirm, 
                 bg="#e74c3c", fg="white", font=("Arial", 11, "bold"), 
                 width=18, height=2).pack(side="left", padx=5)
        tk.Button(btn_frame, text="❌ Cancel", command=on_cancel, 
                 bg="#95a5a6", fg="white", font=("Arial", 11), 
                 width=12, height=2).pack(side="left", padx=5)
        
        dialog.wait_window()
        
        if not result["choice"]:
            return
        
        try:
            self.status_var.set("Clearing data...")
            self.root.update()
            
            # גיבוי
            backup_path = self.backup_file()
            
            wb = load_workbook(SYSTEM_FILE)
            
            # בדיקה איזה גיליון קיים
            periods_sheet = self.get_tracking_sheet_name(wb)
            
            # מחיקת תקופות מילואים
            ws_periods = wb[periods_sheet]
            for row in range(ws_periods.max_row, 1, -1):
                ws_periods.delete_rows(row)
            
            # מחיקת תשלומי ב"ל
            if '3️⃣ תשלומי ב"ל' in wb.sheetnames:
                ws_btl = wb['3️⃣ תשלומי ב"ל']
                for row in range(ws_btl.max_row, 1, -1):
                    ws_btl.delete_rows(row)
            
            # מחיקת רשימת תשלומים
            if '💵 רשימת תשלומים' in wb.sheetnames:
                ws_payments = wb['💵 רשימת תשלומים']
                for row in range(ws_payments.max_row, 1, -1):
                    ws_payments.delete_rows(row)
            
            # מחיקת דוח מסכם (אם קיים)
            if '4️⃣ דוח מסכם' in wb.sheetnames:
                ws_summary = wb['4️⃣ דוח מסכם']
                for row in range(ws_summary.max_row, 1, -1):
                    ws_summary.delete_rows(row)
            
            wb.save(SYSTEM_FILE)
            
            self.status_var.set("All data cleared")
            messagebox.showinfo("Success", 
                f"All data has been deleted!\n\n"
                f"✅ Backup saved to:\n{backup_path}\n\n"
                f"You can now import fresh data.")
            
        except Exception as e:
            self.status_var.set("Error")
            messagebox.showerror("Error", f"Error clearing data:\n{str(e)}")

def main():
    root = tk.Tk()
    app = MiluimManager(root)
    root.mainloop()

if __name__ == "__main__":
    main()
