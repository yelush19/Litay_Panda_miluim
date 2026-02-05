#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
תיקון כותרות דוח מסכם
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os

SYSTEM_FILE = r"C:\Projects\LitayPandaMiluim\מערכת_מילואים_מלאה.xlsx"

# צבעי ליטאי
GREEN_HEADER = PatternFill(start_color="528163", end_color="528163", fill_type="solid")
header_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")

def fix_summary_headers():
    """תיקון כותרות דוח מסכם"""
    
    print("=" * 60)
    print("🔧 תיקון כותרות דוח מסכם")
    print("=" * 60)
    
    if not os.path.exists(SYSTEM_FILE):
        print(f"\n❌ לא נמצא קובץ: {SYSTEM_FILE}")
        input("\nלחץ Enter לסגירה...")
        return
    
    wb = load_workbook(SYSTEM_FILE)
    
    if '4️⃣ דוח מסכם' not in wb.sheetnames:
        print("\n❌ לא נמצא גיליון 'דוח מסכם'")
        wb.close()
        input("\nלחץ Enter לסגירה...")
        return
    
    ws = wb['4️⃣ דוח מסכם']
    
    # הכותרות הנכונות (15 עמודות)
    headers = [
        "שם עובד",
        "מזהה תקופה",
        "מחלקה",
        "חודש",
        "תאריך התחלה",
        "תאריך סיום",
        "סה\"כ ימים",
        "ימי א-ה",
        "תעריף יומי",
        "תשלום מעסיק",
        "תגמול ב\"ל",
        "פיצוי 20%",
        "תוספת 40%",
        "הפרש",
        "סטטוס"
    ]
    
    print(f"\n📝 מעדכן כותרות ({len(headers)} עמודות)...")
    
    # עדכון כותרות
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col)
        cell.value = header
        cell.font = header_font
        cell.fill = GREEN_HEADER
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        print(f"   {col}. {header}")
    
    # מחיקת עמודות מיותרות (16 ואילך)
    if ws.max_column > 15:
        print(f"\n🗑️  מוחק עמודות מיותרות (16-{ws.max_column})...")
        for col in range(ws.max_column, 15, -1):
            ws.delete_cols(col)
    
    # הגדרת RTL
    ws.sheet_view.rightToLeft = True
    
    # שמירה
    wb.save(SYSTEM_FILE)
    wb.close()
    
    print("\n" + "=" * 60)
    print("✅ כותרות דוח מסכם תוקנו בהצלחה!")
    print("=" * 60)
    print("\n📊 מבנה חדש:")
    print("   1-2: מזהים (עובד, תקופה)")
    print("   3-6: מיקום זמן (מחלקה, חודש, תאריכים)")
    print("   7-9: ימים ותעריף")
    print("   10-13: תשלומים (מעסיק, ב\"ל, פיצוי, תוספת)")
    print("   14-15: סיכום (הפרש, סטטוס)")
    
    input("\nלחץ Enter לסגירה...")

if __name__ == "__main__":
    try:
        fix_summary_headers()
    except Exception as e:
        print(f"\n❌ שגיאה: {e}")
        import traceback
        traceback.print_exc()
        input("\nלחץ Enter לסגירה...")
