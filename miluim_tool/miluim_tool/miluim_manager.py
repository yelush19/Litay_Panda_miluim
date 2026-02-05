#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
מערכת ניהול תשלומי מילואים - ליטאי ניהול שירותים
"""

import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import os
import shutil

# === הגדרות ===
SYSTEM_FILE = r"C:\Projects\LitayPandaMiluim\מערכת_מילואים_מלאה.xlsx"

# צבעי ליטאי
LITAY_GREEN = "#528163"
LITAY_GREEN_DARK = "#2d5f3f"
LITAY_GREEN_LIGHT = "#8dd1bb"
LITAY_BG = "#f5f6fa"

class MiluimManager:
    def __init__(self, root):
        self.root = root
        self.root.title("Miluim System - Litay")
        self.root.geometry("500x450")
        self.root.configure(bg=LITAY_BG)
        
        title = tk.Label(root, text="מערכת ניהול תשלומי מילואים",
                        font=("Arial", 18, "bold"), bg=LITAY_BG, fg=LITAY_GREEN_DARK)
        title.pack(pady=20)
        
        subtitle = tk.Label(root, text="Litay - ליטאי ניהול שירותים",
                           font=("Arial", 12), bg=LITAY_BG, fg=LITAY_GREEN)
        subtitle.pack()
        
        btn_frame = tk.Frame(root, bg=LITAY_BG)
        btn_frame.pack(pady=30, padx=40, fill="both", expand=True)
        
        self.create_button(btn_frame, "Import MECANO / ייבוא מקאנו", self.import_mecano)
        self.create_button(btn_frame, "Import BTL Payment / ייבוא תשלום ב״ל", self.import_btl)
        self.create_button(btn_frame, "Import 40% Bonus / ייבוא תוספת 40%", self.import_40_percent)
        self.create_button(btn_frame, "Calculate Differences / חישוב הפרשים", self.calculate_differences)
        
        self.status_var = tk.StringVar(value="Ready / מוכן לעבודה")
        status = tk.Label(root, textvariable=self.status_var, font=("Arial", 10),
                         bg=LITAY_GREEN_LIGHT, fg=LITAY_GREEN_DARK, pady=10)
        status.pack(fill="x", side="bottom")
        
        self.update_all = None
        
    def create_button(self, parent, text, command):
        btn = tk.Button(parent, text=text, font=("Arial", 12), bg=LITAY_GREEN, fg="white",
                       activebackground=LITAY_GREEN_DARK, activeforeground="white",
                       cursor="hand2", command=command, height=2)
        btn.pack(fill="x", pady=8)
    
    def format_date(self, date_val):
        if pd.isna(date_val):
            return ""
        date_str = str(date_val).strip()
        if len(date_str) == 8 and date_str[2] == '/' and date_str[5] == '/':
            parts = date_str.split('/')
            year = '20' + parts[2] if int(parts[2]) < 50 else '19' + parts[2]
            return f"{parts[0]}/{parts[1]}/{year}"
        if hasattr(date_val, 'strftime'):
            return date_val.strftime('%d/%m/%Y')
        return date_str
    
    def normalize_date(self, date_val):
        formatted = self.format_date(date_val)
        return formatted.strip() if formatted else ""
        
    def backup_file(self):
        if os.path.exists(SYSTEM_FILE):
            backup_dir = os.path.join(os.path.dirname(SYSTEM_FILE), "backups")
            os.makedirs(backup_dir, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = os.path.join(backup_dir, f"backup_{timestamp}.xlsx")
            shutil.copy2(SYSTEM_FILE, backup_path)
            return backup_path
        return None
    
    def get_existing_btl_records(self, ws):
        existing = {}
        for row in range(2, ws.max_row + 1):
            tz = str(ws.cell(row, 1).value or "").strip()
            start_date = self.normalize_date(ws.cell(row, 3).value)
            end_date = self.normalize_date(ws.cell(row, 4).value)
            claim_type = str(ws.cell(row, 5).value or "").strip()
            tagmul = ws.cell(row, 6).value or 0
            if tz:
                key = f"{tz}|{start_date}|{end_date}|{claim_type}"
                existing[key] = {"row": row, "tagmul": tagmul}
        return existing
    
    def ask_update_or_skip(self, employee_name, date_start, existing_amount, new_amount):
        if self.update_all is not None:
            return self.update_all
        
        dialog = tk.Toplevel(self.root)
        dialog.title("Duplicate Found")
        dialog.geometry("450x250")
        dialog.configure(bg=LITAY_BG)
        dialog.transient(self.root)
        dialog.grab_set()
        
        result = {"choice": "skip"}
        
        msg = f"""Found existing row with different amount:

Employee: {employee_name}
Date: {date_start}
Existing: {existing_amount:,.0f} NIS
New: {new_amount:,.0f} NIS

What to do?"""
        
        label = tk.Label(dialog, text=msg, font=("Arial", 11), bg=LITAY_BG, 
                        fg=LITAY_GREEN_DARK, justify="left")
        label.pack(pady=20, padx=20)
        
        btn_frame = tk.Frame(dialog, bg=LITAY_BG)
        btn_frame.pack(pady=10)
        
        def on_update():
            result["choice"] = "update"
            dialog.destroy()
        def on_skip():
            result["choice"] = "skip"
            dialog.destroy()
        def on_update_all():
            result["choice"] = "update"
            self.update_all = "update"
            dialog.destroy()
        def on_skip_all():
            result["choice"] = "skip"
            self.update_all = "skip"
            dialog.destroy()
        
        tk.Button(btn_frame, text="Update", font=("Arial", 10), bg=LITAY_GREEN, fg="white",
                 command=on_update, width=10).grid(row=0, column=0, padx=5, pady=5)
        tk.Button(btn_frame, text="Skip", font=("Arial", 10), bg="#999", fg="white",
                 command=on_skip, width=10).grid(row=0, column=1, padx=5, pady=5)
        tk.Button(btn_frame, text="Update All", font=("Arial", 10), bg=LITAY_GREEN_DARK, fg="white",
                 command=on_update_all, width=10).grid(row=1, column=0, padx=5, pady=5)
        tk.Button(btn_frame, text="Skip All", font=("Arial", 10), bg="#666", fg="white",
                 command=on_skip_all, width=10).grid(row=1, column=1, padx=5, pady=5)
        
        dialog.wait_window()
        return result["choice"]
        
    def import_mecano(self):
        file_path = filedialog.askopenfilename(title="Select MECANO file",
                                               filetypes=[("Excel files", "*.xlsx *.xls")])
        if not file_path:
            return
        try:
            self.status_var.set("Importing MECANO...")
            self.root.update()
            df = pd.read_excel(file_path)
            self.backup_file()
            self.status_var.set(f"Imported {len(df)} records from MECANO")
            messagebox.showinfo("Success", f"Imported {len(df)} records from MECANO")
        except Exception as e:
            self.status_var.set("Error importing")
            messagebox.showerror("Error", f"Error:\n{str(e)}")
            
    def import_btl(self):
        file_path = filedialog.askopenfilename(title="Select BTL file",
                                               filetypes=[("Excel files", "*.xlsx *.xls *.xla")])
        if not file_path:
            return
        try:
            self.status_var.set("Importing BTL file...")
            self.root.update()
            
            self.update_all = None
            
            df = pd.read_excel(file_path, header=None)
            
            mana_number = df.iloc[2, 1]
            payment_date = df.iloc[9, 1]
            
            headers = df.iloc[11].tolist()
            data = df.iloc[12:].copy()
            data.columns = headers
            data = data.dropna(subset=['זהות'])
            
            self.backup_file()
            
            wb = load_workbook(SYSTEM_FILE)
            ws = wb['3️⃣ תשלומי ב"ל']
            
            existing = self.get_existing_btl_records(ws)
            
            next_row = ws.max_row + 1
            
            added = 0
            skipped = 0
            updated = 0
            total_tagmul = 0
            total_pitzuy = 0
            
            for _, row in data.iterrows():
                try:
                    tz = str(row['זהות']).strip()
                    employee_name = f"{row['שם פרטי']} {row['שם משפחה']}"
                    start_date = self.normalize_date(row['תאריך שרות'])
                    end_date = self.normalize_date(row['תאריך סיום שרות'])
                    
                    # סוג תביעה - מעמודה "סוג תביעה" (מקור/הפרש)
                    claim_type = str(row['סוג תביעה']).strip()
                    
                    tagmul_raw = row['תגמול']
                    pitzuy_raw = row['פיצוי %20 למעסיק']
                    
                    tagmul = 0
                    if pd.notna(tagmul_raw):
                        tagmul_str = str(tagmul_raw).replace('+', '').replace('-', '').replace(',', '')
                        if tagmul_str and not str(tagmul_raw).startswith('-'):
                            tagmul = float(tagmul_str) if tagmul_str else 0
                    
                    pitzuy = 0
                    if pd.notna(pitzuy_raw):
                        pitzuy_str = str(pitzuy_raw).replace('+', '').replace('-', '').replace(',', '')
                        if pitzuy_str and not str(pitzuy_raw).startswith('-'):
                            pitzuy = float(pitzuy_str) if pitzuy_str else 0
                    
                    key = f"{tz}|{start_date}|{end_date}|{claim_type}"
                    
                    if key in existing:
                        existing_tagmul = existing[key]["tagmul"] or 0
                        existing_row = existing[key]["row"]
                        
                        if abs(existing_tagmul - tagmul) < 1:
                            skipped += 1
                            continue
                        else:
                            choice = self.ask_update_or_skip(employee_name, start_date, 
                                                            existing_tagmul, tagmul)
                            if choice == "skip":
                                skipped += 1
                                continue
                            else:
                                ws.cell(existing_row, 6).value = tagmul
                                ws.cell(existing_row, 7).value = pitzuy
                                ws.cell(existing_row, 9).value = tagmul
                                ws.cell(existing_row, 10).value = mana_number
                                ws.cell(existing_row, 11).value = self.format_date(payment_date)
                                updated += 1
                                total_tagmul += tagmul
                                total_pitzuy += pitzuy
                                continue
                    
                    ws.cell(next_row, 1).value = tz
                    ws.cell(next_row, 2).value = employee_name
                    ws.cell(next_row, 3).value = start_date
                    ws.cell(next_row, 4).value = end_date
                    ws.cell(next_row, 5).value = claim_type  # מקור / הפרש
                    ws.cell(next_row, 6).value = tagmul
                    ws.cell(next_row, 7).value = pitzuy
                    ws.cell(next_row, 8).value = 0
                    ws.cell(next_row, 9).value = tagmul
                    ws.cell(next_row, 10).value = mana_number
                    ws.cell(next_row, 11).value = self.format_date(payment_date)
                    ws.cell(next_row, 12).value = os.path.basename(file_path)
                    
                    total_tagmul += tagmul
                    total_pitzuy += pitzuy
                    next_row += 1
                    added += 1
                    
                except Exception as e:
                    print(f"Row error: {e}")
                    continue
            
            wb.save(SYSTEM_FILE)
            
            self.status_var.set(f"Done: {added} added, {updated} updated, {skipped} skipped")
            messagebox.showinfo("Success", 
                f"Mana: {mana_number}\n"
                f"Date: {self.format_date(payment_date)}\n\n"
                f"Added: {added} new records\n"
                f"Updated: {updated} records\n"
                f"Skipped: {skipped} duplicates\n\n"
                f"Total Tagmul: {total_tagmul:,.0f} NIS\n"
                f"Total Pitzuy 20%: {total_pitzuy:,.0f} NIS")
            
        except Exception as e:
            self.status_var.set("Error importing")
            messagebox.showerror("Error", f"Error:\n{str(e)}")
            
    def import_40_percent(self):
        messagebox.showinfo("Coming Soon", "This feature will be added soon")
        
    def calculate_differences(self):
        messagebox.showinfo("Coming Soon", "This feature will be added soon")

def main():
    root = tk.Tk()
    app = MiluimManager(root)
    root.mainloop()

if __name__ == "__main__":
    main()
