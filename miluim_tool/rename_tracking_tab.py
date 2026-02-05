#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
שינוי שם טאב: תקופות מילואים → מעקב מילואים ותשלומים
"""

from openpyxl import load_workbook
from datetime import datetime
import shutil
import os

SYSTEM_FILE = r"C:\Projects\LitayPandaMiluim\מערכת_מילואים_מלאה.xlsx"

def rename_tab():
    """שינוי שם הטאב"""
    
    print("=" * 60)
    print("🔄 שינוי שם טאב - תקופות מילואים")
    print("=" * 60)
    
    if not os.path.exists(SYSTEM_FILE):
        print(f"\n❌ לא נמצא קובץ: {SYSTEM_FILE}")
        input("\nלחץ Enter לסגירה...")
        return
    
    # גיבוי
    backup_dir = os.path.join(os.path.dirname(SYSTEM_FILE), "backups")
    os.makedirs(backup_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = os.path.join(backup_dir, f"backup_before_rename_{timestamp}.xlsx")
    shutil.copy2(SYSTEM_FILE, backup_path)
    print(f"\n💾 גיבוי: {os.path.basename(backup_path)}")
    
    # טעינה
    wb = load_workbook(SYSTEM_FILE)
    
    old_name = "2️⃣ תקופות מילואים"
    new_name = "📊 מעקב מילואים ותשלומים"
    
    if old_name in wb.sheetnames:
        ws = wb[old_name]
        ws.title = new_name
        print(f"\n✅ שם הטאב שונה:")
        print(f"   מ: {old_name}")
        print(f"   ל: {new_name}")
    elif new_name in wb.sheetnames:
        print(f"\n✅ הטאב כבר קרוי: {new_name}")
    else:
        print(f"\n❌ לא נמצא טאב: {old_name}")
        wb.close()
        input("\nלחץ Enter לסגירה...")
        return
    
    # עדכון הקוד שמשתמש בשם הישן
    print(f"\n📝 רשימת גיליונות עכשיו:")
    for i, sheet in enumerate(wb.sheetnames, 1):
        marker = "✨" if sheet == new_name else "  "
        print(f"   {marker} {i}. {sheet}")
    
    # שמירה
    wb.save(SYSTEM_FILE)
    wb.close()
    
    print("\n" + "=" * 60)
    print("✅ שינוי שם הטאב הושלם!")
    print("=" * 60)
    
    print("\n⚠️  חשוב!")
    print("   הקוד במערכת עדיין מחפש את השם הישן.")
    print("   יש לעדכן את miluim_manager.py")
    
    input("\nלחץ Enter לסגירה...")

if __name__ == "__main__":
    try:
        rename_tab()
    except Exception as e:
        print(f"\n❌ שגיאה: {e}")
        import traceback
        traceback.print_exc()
        input("\nלחץ Enter לסגירה...")
