#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
יצירת תבנית חדשה למערכת מילואים
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime
import os

def create_template():
    """יצירת תבנית חדשה"""
    
    wb = Workbook()
    
    # צבעי ליטאי
    GREEN_HEADER = PatternFill(start_color="528163", end_color="528163", fill_type="solid")
    GREEN_LIGHT = PatternFill(start_color="8dd1bb", end_color="8dd1bb", fill_type="solid")
    
    # גופנים
    header_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
    normal_font = Font(name='Arial', size=10)
    
    # מסגרת
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # === 1. גיליון רשימות ===
    ws_lists = wb.active
    ws_lists.title = "רשימות"
    ws_lists.sheet_view.rightToLeft = True  # RTL
    
    ws_lists['A1'] = "חודשי תשלום"
    ws_lists['A1'].font = header_font
    ws_lists['A1'].fill = GREEN_HEADER
    
    # רשימת חודשים (12 חודשים קדימה)
    current_date = datetime.now()
    for i in range(24):  # 24 חודשים
        month = ((current_date.month + i - 1) % 12) + 1
        year = current_date.year + ((current_date.month + i - 1) // 12)
        ws_lists[f'A{i+2}'] = f"{month:02d}/{year}"
    
    ws_lists['B1'] = "סטטוסים"
    ws_lists['B1'].font = header_font
    ws_lists['B1'].fill = GREEN_HEADER
    
    ws_lists['B2'] = "ממתין"
    ws_lists['B3'] = "שולם"
    ws_lists['B4'] = "מאוזן"
    ws_lists['B5'] = "לא רלוונטי"
    
    # === 2. רשימת עובדים ===
    ws_emp = wb.create_sheet("1️⃣ רשימת עובדים")
    ws_emp.sheet_view.rightToLeft = True  # RTL
    
    headers_emp = [
        "ת.ז.", "שם פרטי", "שם משפחה", "שם מלא", "מחלקה",
        "משכורת חודשית", "תעריף יומי", "בנק", "מספר חשבון", "סטטוס"
    ]
    
    for col, header in enumerate(headers_emp, 1):
        cell = ws_emp.cell(1, col)
        cell.value = header
        cell.font = header_font
        cell.fill = GREEN_HEADER
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = thin_border
    
    # === 3. מעקב מילואים ותשלומים ===
    ws_track = wb.create_sheet("📊 מעקב מילואים ותשלומים")
    ws_track.sheet_view.rightToLeft = True  # RTL
    
    headers_track = [
        "מזהה תקופה", "שם עובד", "מחלקה", "תאריך התחלה", "תאריך סיום",
        "חודש", "סה\"כ ימים", "ימי א-ה", "ימי שישי", "ימי שבת", "ימי חג",
        "תעריף יומי", "תשלום מעסיק (א-ה)", "פיצוי 20% למעסיק",
        "תגמול ב\"ל ₪", "תוספת 40% ₪", "סה\"כ תגמול מביטוח לאומי ₪",
        "מועד תשלום ביטוח לאומי", "סכום הפרשים לעובד ₪",
        "חודש ביצוע תשלום", "סטטוס", "💰 הערות"
    ]
    
    for col, header in enumerate(headers_track, 1):
        cell = ws_track.cell(1, col)
        cell.value = header
        cell.font = header_font
        cell.fill = GREEN_HEADER
        cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        cell.border = thin_border
        ws_track.column_dimensions[cell.column_letter].width = 15
    
    # Dropdown לחודש ביצוע תשלום (עמודה 20)
    dv_month = DataValidation(type="list", formula1="=רשימות!$A$2:$A$25", allow_blank=True)
    dv_month.error = 'בחר חודש מהרשימה'
    dv_month.errorTitle = 'ערך לא תקין'
    ws_track.add_data_validation(dv_month)
    dv_month.add(f'T2:T1000')
    
    # Dropdown לסטטוס (עמודה 21)
    dv_status = DataValidation(type="list", formula1="=רשימות!$B$2:$B$5", allow_blank=True)
    dv_status.error = 'בחר סטטוס מהרשימה'
    dv_status.errorTitle = 'ערך לא תקין'
    ws_track.add_data_validation(dv_status)
    dv_status.add(f'U2:U1000')
    
    # === 4. תשלומי ב"ל ===
    ws_btl = wb.create_sheet("3️⃣ תשלומי ב\"ל")
    ws_btl.sheet_view.rightToLeft = True  # RTL
    
    headers_btl = [
        "ת.ז.", "שם עובד", "תאריך התחלה", "תאריך סיום", "סוג תשלום",
        "תגמול ₪", "פיצוי 20% ₪", "תוספת 40% ₪", "סה\"כ לעובד ₪",
        "מספר מנה", "תאריך תשלום", "קובץ מקור"
    ]
    
    for col, header in enumerate(headers_btl, 1):
        cell = ws_btl.cell(1, col)
        cell.value = header
        cell.font = header_font
        cell.fill = GREEN_HEADER
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = thin_border
        ws_btl.column_dimensions[cell.column_letter].width = 14
    
    # === 5. רשימת תשלומים ===
    ws_pay = wb.create_sheet("💵 רשימת תשלומים")
    ws_pay.sheet_view.rightToLeft = True  # RTL
    
    headers_pay = [
        "מספר מנה", "תאריך תשלום", "תגמול ₪", "פיצוי 20% ₪",
        "תוספת 40% ₪", "סה\"כ ₪"
    ]
    
    for col, header in enumerate(headers_pay, 1):
        cell = ws_pay.cell(1, col)
        cell.value = header
        cell.font = header_font
        cell.fill = GREEN_HEADER
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        ws_pay.column_dimensions[cell.column_letter].width = 16
    
    # === 6. היסטוריית שכר ===
    ws_hist = wb.create_sheet("📈 היסטוריית שכר")
    ws_hist.sheet_view.rightToLeft = True  # RTL
    
    headers_hist = [
        "שם עובד", "תאריך עדכון", "משכורת חודשית", "תעריף יומי", "סיבת שינוי"
    ]
    
    for col, header in enumerate(headers_hist, 1):
        cell = ws_hist.cell(1, col)
        cell.value = header
        cell.font = header_font
        cell.fill = GREEN_HEADER
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = thin_border
        ws_hist.column_dimensions[cell.column_letter].width = 18
    
    # === 7. הוראות שימוש ===
    ws_help = wb.create_sheet("📖 הוראות שימוש", 0)
    ws_help.sheet_view.rightToLeft = True  # RTL
    
    ws_help['A1'] = "מערכת ניהול תשלומי מילואים - ליטאי"
    ws_help['A1'].font = Font(name='Arial', size=16, bold=True, color="2d5f3f")
    
    instructions = [
        "",
        "🎯 תהליך עבודה:",
        "1. ייבא קובץ מקאנו → יווצרו תקופות מילואים",
        "2. ייבא קובץ ביטוח לאומי → יעודכנו תגמולים",
        "3. לחץ 'חישוב מלא' → יחושבו הפרשים",
        "4. סמן 'חודש ביצוע תשלום' לתקופות ששולמו",
        "5. הפק דוח הפרשים לתקופות שטרם שולמו",
        "",
        "📊 הגיליונות:",
        "• רשימת עובדים - נתוני עובדים ותעריפים",
        "• מעקב מילואים ותשלומים - הגיליון המרכזי",
        "• תשלומי ב\"ל - כל התשלומים מביטוח לאומי",
        "• רשימת תשלומים - סיכום לפי מנה",
        "",
        "🎨 צבעים:",
        "• ירוק - שורה חדשה",
        "• כתום - שורה מעודכנת",
        "",
        "💡 טיפים:",
        "• השתמש בסינון Excel לצפייה בהפרשים ממתינים",
        "• עמודת 'הערות' פתוחה לכתיבה חופשית",
        "• גיבוי אוטומטי נשמר לפני כל פעולה"
    ]
    
    for i, line in enumerate(instructions, 2):
        ws_help[f'A{i}'] = line
        ws_help[f'A{i}'].font = Font(name='Arial', size=11)
        ws_help[f'A{i}'].alignment = Alignment(horizontal='right', vertical='top')
    
    ws_help.column_dimensions['A'].width = 80
    
    # שמירה
    output_path = "מערכת_מילואים_תבנית_חדשה.xlsx"
    wb.save(output_path)
    
    print("=" * 60)
    print("✅ תבנית חדשה נוצרה בהצלחה!")
    print("=" * 60)
    print(f"\n📁 נשמר ב: {output_path}")
    print("\n📋 הגיליונות:")
    print("   1. 📖 הוראות שימוש")
    print("   2. רשימות (חודשים + סטטוסים)")
    print("   3. 1️⃣ רשימת עובדים")
    print("   4. 📊 מעקב מילואים ותשלומים")
    print("   5. 3️⃣ תשלומי ב\"ל")
    print("   6. 💵 רשימת תשלומים")
    print("   7. 📈 היסטוריית שכר")
    
    print("\n✨ תכונות חדשות:")
    print("   • Dropdown לחודש ביצוע תשלום")
    print("   • Dropdown לסטטוס")
    print("   • מועד תשלום ביטוח לאומי")
    print("   • עמודת הערות מורחבת")
    
    input("\nלחץ Enter לסגירה...")

if __name__ == "__main__":
    try:
        create_template()
    except Exception as e:
        print(f"\n❌ שגיאה: {e}")
        input("\nלחץ Enter לסגירה...")
