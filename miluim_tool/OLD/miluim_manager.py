#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
מערכת ניהול תשלומי מילואים - ליטאי ניהול שירותים
"""

import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import os
import shutil

# === הגדרות ===
SYSTEM_FILE = r"C:\Projects\LitayPandaMiluim\מערכת_מילואים_מלאה.xlsx"

# צבעי ליטאי
LITAY_GREEN = "#528163"
LITAY_GREEN_DARK = "#2d5f3f"
LITAY_GREEN_LIGHT = "#8dd1bb"
LITAY_BG = "#f5f6fa"

class MiluimManager:
    def __init__(self, root):
        self.root = root
        self.root.title("Miluim System - Litay")
        self.root.geometry("500x450")
        self.root.configure(bg=LITAY_BG)
        
        title = tk.Label(root, text="מערכת ניהול תשלומי מילואים",
                        font=("Arial", 18, "bold"), bg=LITAY_BG, fg=LITAY_GREEN_DARK)
        title.pack(pady=20)
        
        subtitle = tk.Label(root, text="Litay - ליטאי ניהול שירותים",
                           font=("Arial", 12), bg=LITAY_BG, fg=LITAY_GREEN)
        subtitle.pack()
        
        btn_frame = tk.Frame(root, bg=LITAY_BG)
        btn_frame.pack(pady=30, padx=40, fill="both", expand=True)
        
        self.create_button(btn_frame, "Import MECANO / ייבוא מקאנו", self.import_mecano)
        self.create_button(btn_frame, "Import BTL Payment / ייבוא תשלום ב״ל", self.import_btl)
        self.create_button(btn_frame, "Import 40% Bonus / ייבוא תוספת 40%", self.import_40_percent)
        self.create_button(btn_frame, "Calculate Differences / חישוב הפרשים", self.calculate_differences)
        
        self.status_var = tk.StringVar(value="Ready / מוכן לעבודה")
        status = tk.Label(root, textvariable=self.status_var, font=("Arial", 10),
                         bg=LITAY_GREEN_LIGHT, fg=LITAY_GREEN_DARK, pady=10)
        status.pack(fill="x", side="bottom")
        
    def create_button(self, parent, text, command):
        btn = tk.Button(parent, text=text, font=("Arial", 12), bg=LITAY_GREEN, fg="white",
                       activebackground=LITAY_GREEN_DARK, activeforeground="white",
                       cursor="hand2", command=command, height=2)
        btn.pack(fill="x", pady=8)
    
    def format_date(self, date_val):
        """המרת תאריך לפורמט DD/MM/YYYY"""
        if pd.isna(date_val):
            return ""
        date_str = str(date_val).strip()
        
        # אם בפורמט קצר DD/MM/YY
        if len(date_str) == 8 and date_str[2] == '/' and date_str[5] == '/':
            parts = date_str.split('/')
            year = '20' + parts[2] if int(parts[2]) < 50 else '19' + parts[2]
            return f"{parts[0]}/{parts[1]}/{year}"
        
        # אם זה datetime object
        if hasattr(date_val, 'strftime'):
            return date_val.strftime('%d/%m/%Y')
        
        return date_str
        
    def backup_file(self):
        """יצירת גיבוי"""
        if os.path.exists(SYSTEM_FILE):
            backup_dir = os.path.join(os.path.dirname(SYSTEM_FILE), "backups")
            os.makedirs(backup_dir, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = os.path.join(backup_dir, f"backup_{timestamp}.xlsx")
            shutil.copy2(SYSTEM_FILE, backup_path)
            return backup_path
        return None
        
    def import_mecano(self):
        """ייבוא קובץ מקאנו"""
        file_path = filedialog.askopenfilename(title="Select MECANO file",
                                               filetypes=[("Excel files", "*.xlsx *.xls")])
        if not file_path:
            return
        try:
            self.status_var.set("Importing MECANO...")
            self.root.update()
            df = pd.read_excel(file_path)
            self.backup_file()
            self.status_var.set(f"Imported {len(df)} records from MECANO")
            messagebox.showinfo("Success", f"Imported {len(df)} records from MECANO")
        except Exception as e:
            self.status_var.set("Error importing")
            messagebox.showerror("Error", f"Error:\n{str(e)}")
            
    def import_btl(self):
        """ייבוא תשלום ביטוח לאומי"""
        file_path = filedialog.askopenfilename(title="Select BTL file",
                                               filetypes=[("Excel files", "*.xlsx *.xls *.xla")])
        if not file_path:
            return
        try:
            self.status_var.set("Importing BTL file...")
            self.root.update()
            
            df = pd.read_excel(file_path, header=None)
            
            # פרטי מנה
            mana_number = df.iloc[2, 1]
            payment_date = df.iloc[9, 1]
            
            # נתונים
            headers = df.iloc[11].tolist()
            data = df.iloc[12:].copy()
            data.columns = headers
            data = data.dropna(subset=['זהות'])
            
            self.backup_file()
            
            wb = load_workbook(SYSTEM_FILE)
            ws = wb['3️⃣ תשלומי ב"ל']
            next_row = ws.max_row + 1
            
            added = 0
            total_tagmul = 0
            total_pitzuy = 0
            
            for _, row in data.iterrows():
                try:
                    ws.cell(next_row, 1).value = str(row['זהות'])
                    ws.cell(next_row, 2).value = f"{row['שם פרטי']} {row['שם משפחה']}"
                    ws.cell(next_row, 3).value = self.format_date(row['תאריך שרות'])
                    ws.cell(next_row, 4).value = self.format_date(row['תאריך סיום שרות'])
                    ws.cell(next_row, 5).value = row['סוג תביעה']
                    
                    # תגמול ופיצוי
                    tagmul_raw = row['תגמול']
                    pitzuy_raw = row['פיצוי %20 למעסיק']
                    
                    tagmul = 0
                    if pd.notna(tagmul_raw):
                        tagmul_str = str(tagmul_raw).replace('+', '').replace('-', '').replace(',', '')
                        if tagmul_str and not str(tagmul_raw).startswith('-'):
                            tagmul = float(tagmul_str) if tagmul_str else 0
                    
                    pitzuy = 0
                    if pd.notna(pitzuy_raw):
                        pitzuy_str = str(pitzuy_raw).replace('+', '').replace('-', '').replace(',', '')
                        if pitzuy_str and not str(pitzuy_raw).startswith('-'):
                            pitzuy = float(pitzuy_str) if pitzuy_str else 0
                    
                    ws.cell(next_row, 6).value = tagmul
                    ws.cell(next_row, 7).value = pitzuy
                    ws.cell(next_row, 8).value = 0
                    ws.cell(next_row, 9).value = tagmul
                    ws.cell(next_row, 10).value = mana_number
                    ws.cell(next_row, 11).value = self.format_date(payment_date)
                    ws.cell(next_row, 12).value = os.path.basename(file_path)
                    
                    total_tagmul += tagmul
                    total_pitzuy += pitzuy
                    next_row += 1
                    added += 1
                except Exception as e:
                    print(f"Row error: {e}")
                    continue
            
            wb.save(SYSTEM_FILE)
            
            self.status_var.set(f"Imported {added} records (Mana {mana_number})")
            messagebox.showinfo("Success", 
                f"Imported {added} records\n\n"
                f"Mana: {mana_number}\n"
                f"Date: {self.format_date(payment_date)}\n\n"
                f"Total Tagmul: {total_tagmul:,.0f} NIS\n"
                f"Total Pitzuy 20%: {total_pitzuy:,.0f} NIS\n"
                f"Grand Total: {total_tagmul + total_pitzuy:,.0f} NIS")
            
        except Exception as e:
            self.status_var.set("Error importing")
            messagebox.showerror("Error", f"Error:\n{str(e)}")
            
    def import_40_percent(self):
        """ייבוא תוספת 40%"""
        messagebox.showinfo("Coming Soon", "This feature will be added soon")
        
    def calculate_differences(self):
        """חישוב הפרשים"""
        messagebox.showinfo("Coming Soon", "This feature will be added soon")

def main():
    root = tk.Tk()
    app = MiluimManager(root)
    root.mainloop()

if __name__ == "__main__":
    main()
