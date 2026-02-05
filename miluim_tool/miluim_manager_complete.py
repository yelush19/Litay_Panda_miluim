#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
מערכת ניהול תשלומי מילואים מלאה - ליטאי ניהול שירותים
Complete System v2.0
"""

import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, timedelta
import os
import shutil
import calendar

SYSTEM_FILE = r"C:\Projects\LitayPandaMiluim\מערכת_מילואים_מלאה.xlsx"

# צבעי ליטאי
LITAY_GREEN = "#528163"
LITAY_GREEN_DARK = "#2d5f3f"
LITAY_GREEN_LIGHT = "#8dd1bb"
LITAY_BG = "#f5f6fa"

# חגים יהודיים 2025 (תאריכים גרגוריאניים)
JEWISH_HOLIDAYS_2025 = [
    datetime(2025, 4, 13), datetime(2025, 4, 14), datetime(2025, 4, 19), datetime(2025, 4, 20),  # פסח
    datetime(2025, 6, 2), datetime(2025, 6, 3),  # שבועות
    datetime(2025, 9, 23), datetime(2025, 9, 24),  # ראש השנה
    datetime(2025, 10, 2),  # יום כיפור
    datetime(2025, 10, 7), datetime(2025, 10, 8), datetime(2025, 10, 13), datetime(2025, 10, 14),  # סוכות
]

class MiluimManager:
    def __init__(self, root):
        self.root = root
        self.root.title("Miluim System - Litay")
        self.root.geometry("520x520")
        self.root.configure(bg=LITAY_BG)
        
        title = tk.Label(root, text="מערכת ניהול תשלומי מילואים",
                        font=("Arial", 18, "bold"), bg=LITAY_BG, fg=LITAY_GREEN_DARK)
        title.pack(pady=15)
        
        subtitle = tk.Label(root, text="Litay Management Services",
                           font=("Arial", 11), bg=LITAY_BG, fg=LITAY_GREEN)
        subtitle.pack()
        
        btn_frame = tk.Frame(root, bg=LITAY_BG)
        btn_frame.pack(pady=20, padx=40, fill="both", expand=True)
        
        self.create_button(btn_frame, "📥 Import MECANO / ייבוא מקאנו", self.import_mecano)
        self.create_button(btn_frame, "💰 Import BTL Payment / ייבוא תשלום ב״ל", self.import_btl)
        self.create_button(btn_frame, "➕ Import 40% Bonus / ייבוא תוספת 40%", self.import_40_percent)
        self.create_button(btn_frame, "🔄 Calculate All / חישוב מלא", self.calculate_all)
        
        self.status_var = tk.StringVar(value="Ready / מוכן לעבודה")
        status = tk.Label(root, textvariable=self.status_var, font=("Arial", 10),
                         bg=LITAY_GREEN_LIGHT, fg=LITAY_GREEN_DARK, pady=10)
        status.pack(fill="x", side="bottom")
        
        self.update_all = None
        
    def create_button(self, parent, text, command):
        btn = tk.Button(parent, text=text, font=("Arial", 11), bg=LITAY_GREEN, fg="white",
                       activebackground=LITAY_GREEN_DARK, activeforeground="white",
                       cursor="hand2", command=command, height=2)
        btn.pack(fill="x", pady=6)
    
    def normalize_name(self, name):
        if pd.isna(name):
            return ""
        return ' '.join(str(name).strip().split())
    
    def format_date(self, date_val):
        if pd.isna(date_val):
            return ""
        if isinstance(date_val, datetime):
            return date_val.strftime('%d/%m/%Y')
        date_str = str(date_val).strip()
        if len(date_str) == 10 and date_str[2] == '.' and date_str[5] == '.':
            parts = date_str.split('.')
            return f"{parts[0]}/{parts[1]}/{parts[2]}"
        if len(date_str) == 8 and date_str[2] == '/' and date_str[5] == '/':
            parts = date_str.split('/')
            year = '20' + parts[2] if int(parts[2]) < 50 else '19' + parts[2]
            return f"{parts[0]}/{parts[1]}/{year}"
        if hasattr(date_val, 'strftime'):
            return date_val.strftime('%d/%m/%Y')
        return date_str
    
    def normalize_date(self, date_val):
        formatted = self.format_date(date_val)
        return formatted.strip() if formatted else ""
    
    def parse_date(self, date_str):
        """המרת תאריך string לdatetime"""
        if pd.isna(date_str) or not date_str:
            return None
        if isinstance(date_str, datetime):
            return date_str
        try:
            parts = str(date_str).split('/')
            if len(parts) == 3:
                return datetime(int(parts[2]), int(parts[1]), int(parts[0]))
        except:
            pass
        return None
        
    def backup_file(self):
        if os.path.exists(SYSTEM_FILE):
            backup_dir = os.path.join(os.path.dirname(SYSTEM_FILE), "backups")
            os.makedirs(backup_dir, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = os.path.join(backup_dir, f"backup_{timestamp}.xlsx")
            shutil.copy2(SYSTEM_FILE, backup_path)
            return backup_path
        return None
    
    def count_work_days(self, start_date, end_date):
        """חישוב ימי עבודה (א-ה), שישי, שבת, חגים"""
        if not start_date or not end_date:
            return 0, 0, 0, 0
        
        weekdays = 0
        fridays = 0
        saturdays = 0
        holidays = 0
        
        current = start_date
        while current <= end_date:
            weekday = current.weekday()
            if current in JEWISH_HOLIDAYS_2025:
                holidays += 1
            elif weekday == 5:  # שבת
                saturdays += 1
            elif weekday == 4:  # שישי
                fridays += 1
            else:  # א-ה
                weekdays += 1
            current += timedelta(days=1)
        
        return weekdays, fridays, saturdays, holidays
    
    def get_next_period_id(self, ws):
        max_id = 0
        for row in range(2, ws.max_row + 1):
            cell_val = ws.cell(row, 1).value
            if cell_val and str(cell_val).startswith('P'):
                try:
                    num = int(str(cell_val).replace('P', ''))
                    max_id = max(max_id, num)
                except:
                    pass
        return f"P{max_id + 1:04d}"
    
    def ask_name_mapping(self, mecano_name, system_names):
        dialog = tk.Toplevel(self.root)
        dialog.title("Name Mapping")
        dialog.geometry("400x320")
        dialog.configure(bg=LITAY_BG)
        dialog.transient(self.root)
        dialog.grab_set()
        
        result = {"choice": None}
        
        msg = f"Employee from MECANO:\n{mecano_name}\n\nNot found. Map to existing?"
        
        label = tk.Label(dialog, text=msg, font=("Arial", 10), bg=LITAY_BG, 
                        fg=LITAY_GREEN_DARK, justify="left")
        label.pack(pady=15, padx=15)
        
        listbox = tk.Listbox(dialog, font=("Arial", 10), height=7)
        listbox.pack(fill="both", expand=True, padx=15)
        
        for name in sorted(system_names):
            listbox.insert(tk.END, name)
        
        btn_frame = tk.Frame(dialog, bg=LITAY_BG)
        btn_frame.pack(pady=10)
        
        def on_select():
            selection = listbox.curselection()
            if selection:
                result["choice"] = listbox.get(selection[0])
            dialog.destroy()
        
        def on_new():
            result["choice"] = "NEW"
            dialog.destroy()
        
        def on_skip():
            result["choice"] = None
            dialog.destroy()
        
        tk.Button(btn_frame, text="Select", command=on_select, bg=LITAY_GREEN, fg="white", width=9).pack(side="left", padx=3)
        tk.Button(btn_frame, text="New", command=on_new, bg=LITAY_GREEN_DARK, fg="white", width=9).pack(side="left", padx=3)
        tk.Button(btn_frame, text="Skip", command=on_skip, bg="#999", fg="white", width=9).pack(side="left", padx=3)
        
        dialog.wait_window()
        return result["choice"]
    
    def import_mecano(self):
        file_path = filedialog.askopenfilename(title="Select MECANO file",
                                               filetypes=[("Excel files", "*.xlsx *.xls")])
        if not file_path:
            return
        try:
            self.status_var.set("Importing MECANO...")
            self.root.update()
            
            df = pd.read_excel(file_path)
            df['תאריך'] = pd.to_datetime(df['תאריך'], format='%d.%m.%Y')
            df = df.sort_values(['שם עובד', 'תאריך'])
            
            # קיבוץ לתקופות
            periods = []
            current_employee = None
            current_start = None
            current_end = None
            current_dept = None
            current_days = 0
            
            for _, row in df.iterrows():
                employee = self.normalize_name(row['שם עובד'])
                date = row['תאריך']
                dept = row['מחלקה']
                
                if employee != current_employee:
                    if current_employee:
                        periods.append({
                            'עובד': current_employee,
                            'מחלקה': current_dept,
                            'התחלה': current_start,
                            'סיום': current_end,
                            'ימים': current_days
                        })
                    current_employee = employee
                    current_start = date
                    current_end = date
                    current_dept = dept
                    current_days = 1
                else:
                    if (date - current_end).days == 1:
                        current_end = date
                        current_days += 1
                    else:
                        periods.append({
                            'עובד': current_employee,
                            'מחלקה': current_dept,
                            'התחלה': current_start,
                            'סיום': current_end,
                            'ימים': current_days
                        })
                        current_start = date
                        current_end = date
                        current_days = 1
            
            if current_employee:
                periods.append({
                    'עובד': current_employee,
                    'מחלקה': current_dept,
                    'התחלה': current_start,
                    'סיום': current_end,
                    'ימים': current_days
                })
            
            self.backup_file()
            
            wb = load_workbook(SYSTEM_FILE)
            ws_periods = wb['2️⃣ תקופות מילואים']
            ws_employees = wb['1️⃣ רשימת עובדים']
            
            df_employees = pd.read_excel(SYSTEM_FILE, sheet_name='1️⃣ רשימת עובדים')
            system_names = set(df_employees['שם מלא'].dropna().apply(self.normalize_name))
            employee_rates = dict(zip(df_employees['שם מלא'].apply(self.normalize_name), 
                                     df_employees['תעריף יומי']))
            
            existing_periods = {}
            for row in range(2, ws_periods.max_row + 1):
                emp = self.normalize_name(ws_periods.cell(row, 2).value)
                start = self.normalize_date(ws_periods.cell(row, 4).value)
                end = self.normalize_date(ws_periods.cell(row, 5).value)
                key = f"{emp}|{start}|{end}"
                existing_periods[key] = row
            
            added = 0
            skipped = 0
            new_employees = []
            name_mappings = {}
            
            next_row = ws_periods.max_row + 1
            
            for period in periods:
                emp_name = period['עובד']
                
                if emp_name not in system_names and emp_name not in name_mappings:
                    choice = self.ask_name_mapping(emp_name, system_names)
                    if choice == "NEW":
                        new_employees.append(emp_name)
                        system_names.add(emp_name)
                        name_mappings[emp_name] = emp_name
                    elif choice:
                        name_mappings[emp_name] = choice
                    else:
                        skipped += 1
                        continue
                
                final_name = name_mappings.get(emp_name, emp_name)
                
                start_str = self.format_date(period['התחלה'])
                end_str = self.format_date(period['סיום'])
                key = f"{final_name}|{start_str}|{end_str}"
                
                if key in existing_periods:
                    skipped += 1
                    continue
                
                # חישוב ימי עבודה
                weekdays, fridays, saturdays, holidays = self.count_work_days(
                    period['התחלה'], period['סיום'])
                
                period_id = self.get_next_period_id(ws_periods)
                ws_periods.cell(next_row, 1).value = period_id
                ws_periods.cell(next_row, 2).value = final_name
                ws_periods.cell(next_row, 3).value = period['מחלקה']
                ws_periods.cell(next_row, 4).value = start_str
                ws_periods.cell(next_row, 5).value = end_str
                ws_periods.cell(next_row, 6).value = period['התחלה'].strftime('%m/%Y')
                ws_periods.cell(next_row, 7).value = period['ימים']
                ws_periods.cell(next_row, 8).value = weekdays
                ws_periods.cell(next_row, 9).value = fridays
                ws_periods.cell(next_row, 10).value = saturdays
                ws_periods.cell(next_row, 11).value = holidays
                
                # תעריף יומי
                rate = employee_rates.get(final_name, 0)
                ws_periods.cell(next_row, 12).value = rate
                
                # תשלום מעסיק (א-ה בלבד)
                if weekdays > 0:
                    ws_periods.cell(next_row, 13).value = weekdays * rate
                
                next_row += 1
                added += 1
            
            # הוספת עובדים חדשים
            if new_employees:
                next_emp_row = ws_employees.max_row + 1
                for emp_name in new_employees:
                    ws_employees.cell(next_emp_row, 4).value = emp_name
                    ws_employees.cell(next_emp_row, 10).value = "פעיל"
                    next_emp_row += 1
            
            wb.save(SYSTEM_FILE)
            
            self.status_var.set(f"MECANO: {added} added, {skipped} skipped")
            messagebox.showinfo("Success", 
                f"MECANO Import Complete\n\n"
                f"Records: {len(df)}\n"
                f"Periods: {len(periods)}\n\n"
                f"Added: {added}\n"
                f"Skipped: {skipped}\n"
                f"New employees: {len(new_employees)}")
            
        except Exception as e:
            self.status_var.set("Error")
            messagebox.showerror("Error", f"MECANO Error:\n{str(e)}")
    
    def get_existing_btl_records(self, ws):
        existing = {}
        for row in range(2, ws.max_row + 1):
            emp = self.normalize_name(ws.cell(row, 2).value)
            start_date = self.normalize_date(ws.cell(row, 3).value)
            end_date = self.normalize_date(ws.cell(row, 4).value)
            claim_type = str(ws.cell(row, 5).value or "").strip()
            tagmul = ws.cell(row, 6).value or 0
            if emp:
                key = f"{emp}|{start_date}|{end_date}|{claim_type}"
                existing[key] = {"row": row, "tagmul": tagmul}
        return existing
    
    def ask_update_or_skip(self, employee_name, date_start, existing_amount, new_amount):
        if self.update_all is not None:
            return self.update_all
        
        dialog = tk.Toplevel(self.root)
        dialog.title("Duplicate")
        dialog.geometry("420x230")
        dialog.configure(bg=LITAY_BG)
        dialog.transient(self.root)
        dialog.grab_set()
        
        result = {"choice": "skip"}
        
        msg = f"""Found existing with different amount:

{employee_name} | {date_start}
Existing: {existing_amount:,.0f} NIS
New: {new_amount:,.0f} NIS"""
        
        label = tk.Label(dialog, text=msg, font=("Arial", 10), bg=LITAY_BG, 
                        fg=LITAY_GREEN_DARK, justify="left")
        label.pack(pady=15, padx=15)
        
        btn_frame = tk.Frame(dialog, bg=LITAY_BG)
        btn_frame.pack(pady=10)
        
        def on_update():
            result["choice"] = "update"
            dialog.destroy()
        def on_skip():
            result["choice"] = "skip"
            dialog.destroy()
        def on_update_all():
            result["choice"] = "update"
            self.update_all = "update"
            dialog.destroy()
        def on_skip_all():
            result["choice"] = "skip"
            self.update_all = "skip"
            dialog.destroy()
        
        tk.Button(btn_frame, text="Update", font=("Arial", 9), bg=LITAY_GREEN, fg="white",
                 command=on_update, width=9).grid(row=0, column=0, padx=4, pady=4)
        tk.Button(btn_frame, text="Skip", font=("Arial", 9), bg="#999", fg="white",
                 command=on_skip, width=9).grid(row=0, column=1, padx=4, pady=4)
        tk.Button(btn_frame, text="Update All", font=("Arial", 9), bg=LITAY_GREEN_DARK, fg="white",
                 command=on_update_all, width=9).grid(row=1, column=0, padx=4, pady=4)
        tk.Button(btn_frame, text="Skip All", font=("Arial", 9), bg="#666", fg="white",
                 command=on_skip_all, width=9).grid(row=1, column=1, padx=4, pady=4)
        
        dialog.wait_window()
        return result["choice"]
            
    def import_btl(self):
        file_path = filedialog.askopenfilename(title="Select BTL file",
                                               filetypes=[("Excel files", "*.xlsx *.xls *.xla")])
        if not file_path:
            return
        try:
            self.status_var.set("Importing BTL...")
            self.root.update()
            
            self.update_all = None
            
            df = pd.read_excel(file_path, header=None)
            
            mana_number = df.iloc[2, 1]
            payment_date = df.iloc[9, 1]
            
            headers = df.iloc[11].tolist()
            data = df.iloc[12:].copy()
            data.columns = headers
            data = data.dropna(subset=['זהות'])
            
            self.backup_file()
            
            wb = load_workbook(SYSTEM_FILE)
            ws = wb['3️⃣ תשלומי ב"ל']
            ws_payments = wb['💵 רשימת תשלומים']
            
            existing = self.get_existing_btl_records(ws)
            
            next_row = ws.max_row + 1
            
            added = 0
            skipped = 0
            updated = 0
            total_tagmul = 0
            total_pitzuy = 0
            
            for _, row in data.iterrows():
                try:
                    tz = str(row['זהות']).strip()
                    employee_name = self.normalize_name(f"{row['שם פרטי']} {row['שם משפחה']}")
                    start_date = self.normalize_date(row['תאריך שרות'])
                    end_date = self.normalize_date(row['תאריך סיום שרות'])
                    claim_type = str(row['סוג תביעה']).strip()
                    
                    tagmul_raw = row['תגמול']
                    pitzuy_raw = row['פיצוי %20 למעסיק']
                    
                    tagmul = 0
                    if pd.notna(tagmul_raw):
                        tagmul_str = str(tagmul_raw).replace('+', '').replace('-', '').replace(',', '')
                        if tagmul_str and not str(tagmul_raw).startswith('-'):
                            tagmul = float(tagmul_str) if tagmul_str else 0
                    
                    pitzuy = 0
                    if pd.notna(pitzuy_raw):
                        pitzuy_str = str(pitzuy_raw).replace('+', '').replace('-', '').replace(',', '')
                        if pitzuy_str and not str(pitzuy_raw).startswith('-'):
                            pitzuy = float(pitzuy_str) if pitzuy_str else 0
                    
                    key = f"{employee_name}|{start_date}|{end_date}|{claim_type}"
                    
                    if key in existing:
                        existing_tagmul = existing[key]["tagmul"] or 0
                        existing_row = existing[key]["row"]
                        
                        if abs(existing_tagmul - tagmul) < 1:
                            skipped += 1
                            continue
                        else:
                            choice = self.ask_update_or_skip(employee_name, start_date, 
                                                            existing_tagmul, tagmul)
                            if choice == "skip":
                                skipped += 1
                                continue
                            else:
                                ws.cell(existing_row, 6).value = tagmul
                                ws.cell(existing_row, 7).value = pitzuy
                                ws.cell(existing_row, 9).value = tagmul
                                ws.cell(existing_row, 10).value = mana_number
                                ws.cell(existing_row, 11).value = self.format_date(payment_date)
                                updated += 1
                                total_tagmul += tagmul
                                total_pitzuy += pitzuy
                                continue
                    
                    ws.cell(next_row, 1).value = tz
                    ws.cell(next_row, 2).value = employee_name
                    ws.cell(next_row, 3).value = start_date
                    ws.cell(next_row, 4).value = end_date
                    ws.cell(next_row, 5).value = claim_type
                    ws.cell(next_row, 6).value = tagmul
                    ws.cell(next_row, 7).value = pitzuy
                    ws.cell(next_row, 8).value = 0
                    ws.cell(next_row, 9).value = tagmul
                    ws.cell(next_row, 10).value = mana_number
                    ws.cell(next_row, 11).value = self.format_date(payment_date)
                    ws.cell(next_row, 12).value = os.path.basename(file_path)
                    
                    total_tagmul += tagmul
                    total_pitzuy += pitzuy
                    next_row += 1
                    added += 1
                    
                except Exception as e:
                    print(f"Row error: {e}")
                    continue
            
            # עדכון רשימת תשלומים
            mana_exists = False
            for r in range(2, ws_payments.max_row + 1):
                if ws_payments.cell(r, 1).value == mana_number:
                    mana_exists = True
                    ws_payments.cell(r, 3).value = total_tagmul
                    ws_payments.cell(r, 4).value = total_pitzuy
                    ws_payments.cell(r, 6).value = total_tagmul + total_pitzuy
                    break
            
            if not mana_exists:
                next_payment_row = ws_payments.max_row + 1
                ws_payments.cell(next_payment_row, 1).value = mana_number
                ws_payments.cell(next_payment_row, 2).value = self.format_date(payment_date)
                ws_payments.cell(next_payment_row, 3).value = total_tagmul
                ws_payments.cell(next_payment_row, 4).value = total_pitzuy
                ws_payments.cell(next_payment_row, 5).value = 0
                ws_payments.cell(next_payment_row, 6).value = total_tagmul + total_pitzuy
            
            wb.save(SYSTEM_FILE)
            
            self.status_var.set(f"BTL: {added} added, {updated} updated")
            messagebox.showinfo("Success", 
                f"Mana: {mana_number} | {self.format_date(payment_date)}\n\n"
                f"Added: {added} | Updated: {updated} | Skipped: {skipped}\n\n"
                f"Tagmul: {total_tagmul:,.0f}\n"
                f"Pitzuy: {total_pitzuy:,.0f}\n"
                f"Total: {total_tagmul + total_pitzuy:,.0f} NIS")
            
        except Exception as e:
            self.status_var.set("Error")
            messagebox.showerror("Error", f"BTL Error:\n{str(e)}")
            
    def import_40_percent(self):
        """ייבוא תוספת 40% - מבנה זהה לייבוא רגיל"""
        file_path = filedialog.askopenfilename(title="Select 40% Bonus file",
                                               filetypes=[("Excel files", "*.xlsx *.xls *.xla")])
        if not file_path:
            return
        try:
            self.status_var.set("Importing 40%...")
            self.root.update()
            
            self.update_all = None
            
            df = pd.read_excel(file_path, header=None)
            
            mana_number = df.iloc[2, 1]
            payment_date = df.iloc[9, 1]
            
            headers = df.iloc[11].tolist()
            data = df.iloc[12:].copy()
            data.columns = headers
            data = data.dropna(subset=['זהות'])
            
            self.backup_file()
            
            wb = load_workbook(SYSTEM_FILE)
            ws = wb['3️⃣ תשלומי ב"ל']
            ws_payments = wb['💵 רשימת תשלומים']
            
            existing = self.get_existing_btl_records(ws)
            
            next_row = ws.max_row + 1
            
            added = 0
            skipped = 0
            total_40 = 0
            
            for _, row in data.iterrows():
                try:
                    tz = str(row['זהות']).strip()
                    employee_name = self.normalize_name(f"{row['שם פרטי']} {row['שם משפחה']}")
                    start_date = self.normalize_date(row['תאריך שרות'])
                    end_date = self.normalize_date(row['תאריך סיום שרות'])
                    
                    # זיהוי תוספת 40% - יכול להיות "תוספת 40%" או שם אחר
                    claim_type = "תוספת 40%"
                    
                    # תוספת 40% מגיעה בעמודה אחרת או זהה
                    bonus_40_raw = row.get('תגמול נדרש', row.get('תגמול', 0))
                    
                    bonus_40 = 0
                    if pd.notna(bonus_40_raw):
                        bonus_str = str(bonus_40_raw).replace('+', '').replace('-', '').replace(',', '')
                        if bonus_str and not str(bonus_40_raw).startswith('-'):
                            bonus_40 = float(bonus_str) if bonus_str else 0
                    
                    if bonus_40 == 0:
                        continue
                    
                    key = f"{employee_name}|{start_date}|{end_date}|{claim_type}"
                    
                    if key in existing:
                        skipped += 1
                        continue
                    
                    ws.cell(next_row, 1).value = tz
                    ws.cell(next_row, 2).value = employee_name
                    ws.cell(next_row, 3).value = start_date
                    ws.cell(next_row, 4).value = end_date
                    ws.cell(next_row, 5).value = claim_type
                    ws.cell(next_row, 6).value = 0
                    ws.cell(next_row, 7).value = 0
                    ws.cell(next_row, 8).value = bonus_40
                    ws.cell(next_row, 9).value = bonus_40
                    ws.cell(next_row, 10).value = mana_number
                    ws.cell(next_row, 11).value = self.format_date(payment_date)
                    ws.cell(next_row, 12).value = os.path.basename(file_path)
                    
                    total_40 += bonus_40
                    next_row += 1
                    added += 1
                    
                except Exception as e:
                    print(f"Row error: {e}")
                    continue
            
            # עדכון רשימת תשלומים
            for r in range(2, ws_payments.max_row + 1):
                if ws_payments.cell(r, 1).value == mana_number:
                    ws_payments.cell(r, 5).value = total_40
                    current_total = (ws_payments.cell(r, 3).value or 0) + \
                                   (ws_payments.cell(r, 4).value or 0) + total_40
                    ws_payments.cell(r, 6).value = current_total
                    break
            
            wb.save(SYSTEM_FILE)
            
            self.status_var.set(f"40%: {added} added")
            messagebox.showinfo("Success", 
                f"40% Bonus Import\n\n"
                f"Mana: {mana_number}\n"
                f"Added: {added}\n"
                f"Total 40%: {total_40:,.0f} NIS")
            
        except Exception as e:
            self.status_var.set("Error")
            messagebox.showerror("Error", f"40% Error:\n{str(e)}")
        
    def calculate_all(self):
        """חישוב מלא - דוח מסכם והפרשים"""
        try:
            self.status_var.set("Calculating...")
            self.root.update()
            
            self.backup_file()
            
            wb = load_workbook(SYSTEM_FILE)
            ws_periods = wb['2️⃣ תקופות מילואים']
            ws_btl = wb['3️⃣ תשלומי ב"ל']
            ws_summary = wb['4️⃣ דוח מסכם']
            ws_employees = wb['1️⃣ רשימת עובדים']
            
            # קריאת תעריפים ושכר חודשי
            df_employees = pd.read_excel(SYSTEM_FILE, sheet_name='1️⃣ רשימת עובדים')
            employee_data = {}
            for _, emp in df_employees.iterrows():
                name = self.normalize_name(emp['שם מלא'])
                employee_data[name] = {
                    'rate': emp.get('תעריף יומי', 0),
                    'monthly': emp.get('משכורת חודשית', 0)
                }
            
            # קריאת תקופות מילואים
            df_periods = pd.read_excel(SYSTEM_FILE, sheet_name='2️⃣ תקופות מילואים')
            
            # קריאת תשלומי ב"ל
            df_btl = pd.read_excel(SYSTEM_FILE, sheet_name='3️⃣ תשלומי ב"ל')
            
            # קיבוץ לפי עובד + חודש
            summary_data = []
            
            for (emp, month), group in df_periods.groupby([df_periods['שם עובד'].apply(self.normalize_name), 
                                                           df_periods['חודש']]):
                total_days = group['סה"כ ימים'].sum()
                weekdays = group['ימי א-ה'].sum()
                
                # תשלום מעסיק
                emp_info = employee_data.get(emp, {})
                rate = emp_info.get('rate', 0)
                monthly = emp_info.get('monthly', 0)
                
                if weekdays > 20:
                    employer_payment = monthly
                else:
                    employer_payment = weekdays * rate
                
                # משיכת תשלומי ב"ל
                btl_payments = df_btl[
                    (df_btl['שם עובד'].apply(self.normalize_name) == emp) &
                    (df_btl['תאריך התחלה'].apply(lambda x: self.parse_date(x).strftime('%m/%Y') if self.parse_date(x) else '') == month)
                ]
                
                btl_tagmul = btl_payments['תגמול ₪'].sum()
                btl_40 = btl_payments['תוספת 40% ₪'].sum()
                
                difference = employer_payment - btl_tagmul
                
                summary_data.append({
                    'עובד': emp,
                    'חודש': month,
                    'תקופות': len(group),
                    'ימים': total_days,
                    'ימי א-ה': weekdays,
                    'תשלום מעסיק': employer_payment,
                    'תגמול ב"ל': btl_tagmul,
                    'תוספת 40%': btl_40,
                    'הפרש': difference
                })
            
            # ניקוי דוח מסכם
            for row in range(ws_summary.max_row, 1, -1):
                if row > 1:
                    ws_summary.delete_rows(row)
            
            # כתיבת דוח מסכם
            next_row = 2
            for item in summary_data:
                ws_summary.cell(next_row, 1).value = item['עובד']
                ws_summary.cell(next_row, 3).value = item['חודש']
                ws_summary.cell(next_row, 6).value = item['תקופות']
                ws_summary.cell(next_row, 7).value = item['ימים']
                ws_summary.cell(next_row, 8).value = item['ימי א-ה']
                ws_summary.cell(next_row, 10).value = item['תשלום מעסיק']
                ws_summary.cell(next_row, 11).value = item['תגמול ב"ל']
                ws_summary.cell(next_row, 12).value = item['הפרש']
                
                # סטטוס
                if abs(item['הפרש']) < 1:
                    status = "מאוזן"
                elif item['הפרש'] > 0:
                    status = "ממתין"
                else:
                    status = "לא רלוונטי"
                ws_summary.cell(next_row, 13).value = status
                
                next_row += 1
            
            wb.save(SYSTEM_FILE)
            
            total_employer = sum(x['תשלום מעסיק'] for x in summary_data)
            total_btl = sum(x['תגמול ב"ל'] for x in summary_data)
            total_diff = sum(x['הפרש'] for x in summary_data)
            
            self.status_var.set("Calculation complete")
            messagebox.showinfo("Success", 
                f"Calculation Complete\n\n"
                f"Summary: {len(summary_data)} rows\n\n"
                f"Employer: {total_employer:,.0f} NIS\n"
                f"BTL: {total_btl:,.0f} NIS\n"
                f"Difference: {total_diff:,.0f} NIS")
            
        except Exception as e:
            self.status_var.set("Error")
            messagebox.showerror("Error", f"Calculation Error:\n{str(e)}")

def main():
    root = tk.Tk()
    app = MiluimManager(root)
    root.mainloop()

if __name__ == "__main__":
    main()
