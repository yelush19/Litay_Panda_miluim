#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
תיקון נתונים קיימים: רגיל → מקור
"""

from openpyxl import load_workbook
import os

SYSTEM_FILE = r"C:\Projects\LitayPandaMiluim\מערכת_מילואים_מלאה.xlsx"

print("=" * 50)
print("  Fixing existing data: רגיל → מקור")
print("=" * 50)

if not os.path.exists(SYSTEM_FILE):
    print(f"\nError: File not found!\n{SYSTEM_FILE}")
    input("\nPress Enter to exit...")
    exit()

wb = load_workbook(SYSTEM_FILE)
ws = wb['3️⃣ תשלומי ב"ל']

updated = 0
for row in range(2, ws.max_row + 1):
    cell = ws.cell(row, 5)  # עמודה E = סוג תשלום
    if cell.value == "רגיל":
        cell.value = "מקור"
        updated += 1

wb.save(SYSTEM_FILE)

print(f"\nUpdated {updated} rows: רגיל → מקור")
print("\nDone!")
input("\nPress Enter to exit...")
