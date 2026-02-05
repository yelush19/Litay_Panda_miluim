#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ייבוא נתוני תשלומים ועדכוני סטטוס
מעדכן: תוספת 40%, מועד תשלום ב"ל, חודש ביצוע תשלום, הערות
"""

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
import shutil
import os

SYSTEM_FILE = r"C:\Projects\LitayPandaMiluim\מערכת_מילואים_מלאה.xlsx"
IMPORT_FILE = r"C:\Projects\LitayPandaMiluim\ריכוז_תשלומים_ועדכוני_סטטוס_רטרו.xlsx"

# צבעים
COLOR_NEW = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")  # ירוק
COLOR_UPDATED = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")  # כתום

def backup_file():
    """יצירת גיבוי"""
    if os.path.exists(SYSTEM_FILE):
        backup_dir = os.path.join(os.path.dirname(SYSTEM_FILE), "backups")
        os.makedirs(backup_dir, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = os.path.join(backup_dir, f"backup_before_payment_import_{timestamp}.xlsx")
        shutil.copy2(SYSTEM_FILE, backup_path)
        print(f"✅ גיבוי נוצר: {backup_path}")
        return backup_path
    return None

def import_payment_data():
    """ייבוא נתוני תשלומים"""
    
    print("=" * 60)
    print("📥 ייבוא נתוני תשלומים ועדכוני סטטוס")
    print("=" * 60)
    print("\n⚠️  חשוב: לפני הריצה, פתח את הקובץ:")
    print("   ריכוז_תשלומים_ועדכוני_סטטוס_רטרו.xlsx")
    print("   ב-Excel ושמור אותו (כדי לחשב נוסחאות)")
    print()
    
    # גיבוי
    backup_path = backup_file()
    
    # טעינת קבצים - עם data_only כדי לקבל ערכים מחושבים
    print("\n📂 טוען קבצים...")
    wb_system = load_workbook(SYSTEM_FILE)
    wb_import = load_workbook(IMPORT_FILE, data_only=True)  # קריאת ערכים מחושבים
    
    # שם הגיליון במערכת - צריך לבדוק אם זה שם ישן או חדש
    if '📊 מעקב מילואים ותשלומים' in wb_system.sheetnames:
        ws_system = wb_system['📊 מעקב מילואים ותשלומים']
        sheet_name = '📊 מעקב מילואים ותשלומים'
    else:
        ws_system = wb_system['2️⃣ תקופות מילואים']
        sheet_name = '2️⃣ תקופות מילואים'
    
    ws_import = wb_import['גיליון1']
    
    print(f"   ✅ מערכת: {sheet_name}")
    print(f"   ✅ קובץ ייבוא: גיליון1")
    
    # בניית מילון מזהי תקופות במערכת
    print("\n🔍 בודק מזהי תקופות במערכת...")
    system_periods = {}
    for row in range(2, ws_system.max_row + 1):
        period_id = ws_system.cell(row, 1).value
        if period_id:
            system_periods[str(period_id).strip()] = row
    
    print(f"   ✅ נמצאו {len(system_periods)} תקופות במערכת")
    
    # עיבוד נתונים מקובץ הייבוא
    print("\n📊 מעבד נתונים...")
    new_count = 0
    updated_count = 0
    skipped_count = 0
    
    for row in range(2, ws_import.max_row + 1):
        period_id = ws_import.cell(row, 1).value  # עמודה A - מזהה תקופה
        
        if not period_id:
            skipped_count += 1
            continue
        
        period_id = str(period_id).strip()
        
        # נתוני התשלומים מהקובץ
        bonus_40 = ws_import.cell(row, 15).value  # עמודה O - תוספת 40%
        total_btl = ws_import.cell(row, 16).value  # עמודה P - סה"כ תגמול מב"ל
        btl_date = ws_import.cell(row, 17).value  # עמודה Q - מועד תשלום ב"ל
        diff_amount = ws_import.cell(row, 18).value  # עמודה R - הפרשים
        payment_month = ws_import.cell(row, 19).value  # עמודה S - חודש ביצוע תשלום
        notes = ws_import.cell(row, 20).value  # עמודה T - הערות
        
        if period_id in system_periods:
            # עדכון שורה קיימת
            system_row = system_periods[period_id]
            
            # עמודות לעדכון - כולל תגמול ופיצוי!
            ws_system.cell(system_row, 13).value = ws_import.cell(row, 13).value  # תשלום מעסיק
            ws_system.cell(system_row, 14).value = ws_import.cell(row, 14).value  # פיצוי 20%
            ws_system.cell(system_row, 15).value = bonus_40  # תוספת 40%
            ws_system.cell(system_row, 16).value = total_btl  # סה"כ תגמול
            ws_system.cell(system_row, 17).value = btl_date  # מועד תשלום ב"ל
            ws_system.cell(system_row, 18).value = diff_amount  # הפרשים
            ws_system.cell(system_row, 19).value = payment_month  # חודש ביצוע
            ws_system.cell(system_row, 20).value = notes  # הערות
            
            # צביעה בכתום
            for col in range(1, 21):
                ws_system.cell(system_row, col).fill = COLOR_UPDATED
            
            updated_count += 1
            
        else:
            # הוספת שורה חדשה
            new_row = ws_system.max_row + 1
            
            # העתקת כל העמודות
            for col in range(1, 21):
                value = ws_import.cell(row, col).value
                ws_system.cell(new_row, col).value = value
                ws_system.cell(new_row, col).fill = COLOR_NEW
            
            new_count += 1
    
    # שמירה
    wb_system.save(SYSTEM_FILE)
    wb_system.close()
    wb_import.close()
    
    print("\n" + "=" * 60)
    print("✅ ייבוא הושלם בהצלחה!")
    print("=" * 60)
    print(f"\n📊 סיכום:")
    print(f"   🆕 שורות חדשות (ירוק): {new_count}")
    print(f"   🔄 שורות מעודכנות (כתום): {updated_count}")
    print(f"   ⏭️  שורות דלגו: {skipped_count}")
    
    print(f"\n💾 גיבוי נשמר ב:\n   {backup_path}")
    
    print("\n📝 עמודות שעודכנו:")
    print("   • תשלום מעסיק (א-ה)")
    print("   • פיצוי 20% למעסיק")
    print("   • תוספת 40%")
    print("   • סה\"כ תגמול מביטוח לאומי")
    print("   • מועד תשלום ביטוח לאומי")
    print("   • סכום הפרשים לעובד")
    print("   • חודש ביצוע תשלום")
    print("   • הערות")
    
    input("\nלחץ Enter לסגירה...")

if __name__ == "__main__":
    try:
        if not os.path.exists(SYSTEM_FILE):
            print(f"❌ לא נמצא קובץ מערכת:\n   {SYSTEM_FILE}")
            input("\nלחץ Enter לסגירה...")
        elif not os.path.exists(IMPORT_FILE):
            print(f"❌ לא נמצא קובץ ייבוא:\n   {IMPORT_FILE}")
            input("\nלחץ Enter לסגירה...")
        else:
            import_payment_data()
    except Exception as e:
        print(f"\n❌ שגיאה: {e}")
        import traceback
        traceback.print_exc()
        input("\nלחץ Enter לסגירה...")
