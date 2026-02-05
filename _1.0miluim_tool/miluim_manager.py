#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
מערכת ניהול תשלומי מילואים - ליטאי ניהול שירותים
כלי ייבוא ועדכון נתונים
"""

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import os
import shutil

# === הגדרות ===
SYSTEM_FILE = r"C:\Projects\LitayPandaMiluim\מערכת_מילואים_מלאה.xlsx"

# צבעי ליטאי
LITAY_GREEN = "#528163"
LITAY_GREEN_DARK = "#2d5f3f"
LITAY_GREEN_LIGHT = "#8dd1bb"
LITAY_BG = "#f5f6fa"

class MiluimManager:
    def __init__(self, root):
        self.root = root
        self.root.title("Miluim System - Litay")
        self.root.geometry("500x450")
        self.root.configure(bg=LITAY_BG)
        
        # כותרת
        title = tk.Label(
            root, 
            text="מערכת ניהול תשלומי מילואים",
            font=("Arial", 18, "bold"),
            bg=LITAY_BG,
            fg=LITAY_GREEN_DARK
        )
        title.pack(pady=20)
        
        subtitle = tk.Label(
            root,
            text="Litay - ליטאי ניהול שירותים",
            font=("Arial", 12),
            bg=LITAY_BG,
            fg=LITAY_GREEN
        )
        subtitle.pack()
        
        # מסגרת כפתורים
        btn_frame = tk.Frame(root, bg=LITAY_BG)
        btn_frame.pack(pady=30, padx=40, fill="both", expand=True)
        
        # כפתורים
        self.create_button(btn_frame, "Import MECANO / ייבוא מקאנו", self.import_mecano, 0)
        self.create_button(btn_frame, "Import BTL Payment / ייבוא תשלום ב״ל", self.import_btl, 1)
        self.create_button(btn_frame, "Import 40% Bonus / ייבוא תוספת 40%", self.import_40_percent, 2)
        self.create_button(btn_frame, "Calculate Differences / חישוב הפרשים", self.calculate_differences, 3)
        
        # סטטוס
        self.status_var = tk.StringVar(value="Ready / מוכן לעבודה")
        status = tk.Label(
            root,
            textvariable=self.status_var,
            font=("Arial", 10),
            bg=LITAY_GREEN_LIGHT,
            fg=LITAY_GREEN_DARK,
            pady=10
        )
        status.pack(fill="x", side="bottom")
        
    def create_button(self, parent, text, command, row):
        btn = tk.Button(
            parent,
            text=text,
            font=("Arial", 12),
            bg=LITAY_GREEN,
            fg="white",
            activebackground=LITAY_GREEN_DARK,
            activeforeground="white",
            cursor="hand2",
            command=command,
            height=2
        )
        btn.pack(fill="x", pady=8)
        
    def backup_file(self):
        """יצירת גיבוי של הקובץ"""
        if os.path.exists(SYSTEM_FILE):
            backup_dir = os.path.join(os.path.dirname(SYSTEM_FILE), "backups")
            os.makedirs(backup_dir, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_name = f"system_backup_{timestamp}.xlsx"
            backup_path = os.path.join(backup_dir, backup_name)
            shutil.copy2(SYSTEM_FILE, backup_path)
            return backup_path
        return None
        
    def import_mecano(self):
        """ייבוא קובץ מקאנו"""
        file_path = filedialog.askopenfilename(
            title="Select MECANO file",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if not file_path:
            return
            
        try:
            self.status_var.set("Importing MECANO...")
            self.root.update()
            
            # קריאת הקובץ
            df = pd.read_excel(file_path)
            
            # גיבוי
            self.backup_file()
            
            # TODO: לוגיקת ייבוא מקאנו
            
            self.status_var.set(f"Imported {len(df)} records from MECANO")
            messagebox.showinfo("Success", f"Imported {len(df)} records from MECANO")
            
        except Exception as e:
            self.status_var.set("Error importing")
            messagebox.showerror("Error", f"Error importing MECANO:\n{str(e)}")
            
    def import_btl(self):
        """ייבוא תשלום ביטוח לאומי"""
        file_path = filedialog.askopenfilename(
            title="Select BTL file",
            filetypes=[("Excel files", "*.xlsx *.xls *.xla")]
        )
        if not file_path:
            return
            
        try:
            self.status_var.set("Importing BTL file...")
            self.root.update()
            
            # קריאת הקובץ - שורה 11 היא הכותרות
            df = pd.read_excel(file_path, header=None)
            
            # חילוץ פרטי מנה
            mana_number = df.iloc[2, 1]  # מספר מנה
            payment_date = df.iloc[9, 1]  # תאריך תשלום
            
            # חילוץ נתונים (משורה 12)
            headers = df.iloc[11].tolist()
            data = df.iloc[12:].copy()
            data.columns = headers
            data = data.dropna(subset=['זהות'])
            
            # גיבוי
            self.backup_file()
            
            # טעינת קובץ המערכת
            wb = load_workbook(SYSTEM_FILE)
            ws = wb['3️⃣ תשלומי ב"ל']
            
            # מציאת שורה ריקה ראשונה
            next_row = ws.max_row + 1
            
            # הוספת הנתונים
            added = 0
            total_tagmul = 0
            total_pitzuy = 0
            
            for _, row in data.iterrows():
                try:
                    ws.cell(next_row, 1).value = str(row['זהות'])
                    ws.cell(next_row, 2).value = f"{row['שם פרטי']} {row['שם משפחה']}"
                    ws.cell(next_row, 3).value = row['תאריך שרות']
                    ws.cell(next_row, 4).value = row['תאריך סיום שרות']
                    ws.cell(next_row, 5).value = row['סוג תביעה']
                    
                    # תגמול - עמודה 14 (אינדקס 14 = "תגמול")
                    # ניקוי סימני + ו-
                    tagmul_raw = row['תגמול']
                    pitzuy_raw = row['פיצוי %20 למעסיק']
                    
                    # המרה למספר
                    if pd.notna(tagmul_raw):
                        tagmul = float(str(tagmul_raw).replace('+', '').replace('-', '').replace(',', ''))
                        if str(tagmul_raw).startswith('-'):
                            tagmul = 0  # ערך שלילי = 0
                    else:
                        tagmul = 0
                        
                    if pd.notna(pitzuy_raw):
                        pitzuy = float(str(pitzuy_raw).replace('+', '').replace('-', '').replace(',', ''))
                        if str(pitzuy_raw).startswith('-'):
                            pitzuy = 0
                    else:
                        pitzuy = 0
                    
                    ws.cell(next_row, 6).value = tagmul  # תגמול
                    ws.cell(next_row, 7).value = pitzuy  # פיצוי 20%
                    ws.cell(next_row, 8).value = 0  # תוספת 40%
                    ws.cell(next_row, 9).value = tagmul  # סה"כ לעובד (רק תגמול!)
                    ws.cell(next_row, 10).value = mana_number
                    ws.cell(next_row, 11).value = payment_date
                    ws.cell(next_row, 12).value = os.path.basename(file_path)
                    
                    total_tagmul += tagmul
                    total_pitzuy += pitzuy
                    next_row += 1
                    added += 1
                except Exception as e:
                    print(f"Error in row: {e}")
                    continue
            
            # שמירה
            wb.save(SYSTEM_FILE)
            
            self.status_var.set(f"Imported {added} records (Mana {mana_number})")
            messagebox.showinfo("Success", 
                f"Imported {added} records\n\n"
                f"Mana: {mana_number}\n"
                f"Date: {payment_date}\n\n"
                f"Total Tagmul: {total_tagmul:,.0f} NIS\n"
                f"Total Pitzuy 20%: {total_pitzuy:,.0f} NIS\n"
                f"Grand Total: {total_tagmul + total_pitzuy:,.0f} NIS")
            
        except Exception as e:
            self.status_var.set("Error importing")
            messagebox.showerror("Error", f"Error:\n{str(e)}")
            
    def import_40_percent(self):
        """ייבוא תוספת 40%"""
        file_path = filedialog.askopenfilename(
            title="Select 40% Bonus file",
            filetypes=[("Excel files", "*.xlsx *.xls *.xla")]
        )
        if not file_path:
            return
            
        try:
            self.status_var.set("Importing 40% bonus...")
            self.root.update()
            
            # TODO: לוגיקת ייבוא תוספת 40%
            
            messagebox.showinfo("Coming Soon", "This feature will be added soon")
            self.status_var.set("Ready / מוכן לעבודה")
            
        except Exception as e:
            self.status_var.set("Error importing")
            messagebox.showerror("Error", f"Error:\n{str(e)}")
            
    def calculate_differences(self):
        """חישוב הפרשים ועדכון סטטוסים"""
        try:
            self.status_var.set("Calculating differences...")
            self.root.update()
            
            # גיבוי
            self.backup_file()
            
            # TODO: לוגיקת חישוב הפרשים
            
            messagebox.showinfo("Coming Soon", "This feature will be added soon")
            self.status_var.set("Ready / מוכן לעבודה")
            
        except Exception as e:
            self.status_var.set("Error calculating")
            messagebox.showerror("Error", f"Error:\n{str(e)}")

def main():
    root = tk.Tk()
    app = MiluimManager(root)
    root.mainloop()

if __name__ == "__main__":
    main()
